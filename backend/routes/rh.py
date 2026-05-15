"""
RH Routes - Human Resources module endpoints
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Body, Form, Response
from fastapi.responses import FileResponse
from typing import Optional, List
from datetime import datetime, timezone, timedelta
import uuid
import os
import io
import json
import shutil
import logging
import threading
from pathlib import Path

import httpx

logger = logging.getLogger(__name__)

from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv

# Load environment
ROOT_DIR = Path(__file__).parent.parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Collections
funcionarios_collection = db["funcionarios"]
ponto_collection = db["ponto_registros"]
ponto_abonos_collection = db["ponto_abonos"]
ponto_observacoes_collection = db["ponto_observacoes"]
jornadas_collection = db["jornadas_trabalho"]
custos_rh_config_collection = db["custos_rh_config"]
folha_pagamento_collection = db["folha_pagamento"]
ferias_collection = db["ferias"]
epi_fichas_collection = db["epi_fichas"]
epi_cargos_collection = db["epi_cargos"]

# Create router
rh_router = APIRouter(prefix="/rh", tags=["RH"])

# ===== Jornada de trabalho padrão =====
JORNADA_PADRAO = {
    "seg_sex": {
        "entrada": "08:00",
        "saida_almoco": "11:30",
        "retorno_almoco": "13:30",
        "saida": "18:00",
        "horas_diarias": 8.0
    },
    "sabado": {
        "entrada": "08:00",
        "saida": "12:00",
        "horas_diarias": 4.0
    }
}

# ===== Tabelas de alíquotas 2025 =====
TABELA_INSS_2025 = [
    {"ate": 1518.00, "aliquota": 7.5},
    {"ate": 2793.88, "aliquota": 9.0},
    {"ate": 4190.83, "aliquota": 12.0},
    {"ate": 8157.41, "aliquota": 14.0}
]
TETO_INSS = 951.01

TABELA_IRPF_2025 = [
    {"ate": 2259.20, "aliquota": 0, "deducao": 0},
    {"ate": 2826.65, "aliquota": 7.5, "deducao": 169.44},
    {"ate": 3751.05, "aliquota": 15.0, "deducao": 381.44},
    {"ate": 4664.68, "aliquota": 22.5, "deducao": 662.77},
    {"ate": 999999999.99, "aliquota": 27.5, "deducao": 896.00}
]

FGTS_ALIQUOTA = 8.0
INSS_PATRONAL_ALIQUOTA = 20.0


# ===== Helper Functions =====
def calcular_inss(salario_bruto: float) -> float:
    inss = 0.0
    salario_restante = salario_bruto
    faixa_anterior = 0.0
    
    for faixa in TABELA_INSS_2025:
        if salario_restante <= 0:
            break
        base = min(salario_restante, faixa["ate"] - faixa_anterior)
        inss += base * (faixa["aliquota"] / 100)
        salario_restante -= base
        faixa_anterior = faixa["ate"]
    
    return min(inss, TETO_INSS)


def calcular_irpf(base_calculo: float) -> float:
    for faixa in TABELA_IRPF_2025:
        if base_calculo <= faixa["ate"]:
            return max(0, (base_calculo * faixa["aliquota"] / 100) - faixa["deducao"])
    return 0.0


# ===== PYDANTIC MODELS =====
from pydantic import BaseModel

class FuncionarioCreate(BaseModel):
    nome: str
    cpf: Optional[str] = None
    rg: Optional[str] = None
    data_nascimento: Optional[str] = None
    telefone: Optional[str] = None
    celular: Optional[str] = None
    email: Optional[str] = None
    cep: Optional[str] = None
    endereco: Optional[str] = None
    numero: Optional[str] = None
    complemento: Optional[str] = None
    bairro: Optional[str] = None
    cidade: Optional[str] = None
    uf: Optional[str] = None
    cargo: str
    funcao: Optional[str] = None
    departamento: Optional[str] = None
    salario: float = 0
    data_admissao: str
    regime_contratacao: str = "CLT"
    status: str = "ativo"
    observacoes: Optional[str] = None


class PontoCreate(BaseModel):
    funcionario_id: str
    data: str
    entrada: str
    saida_almoco: Optional[str] = ""
    retorno_almoco: Optional[str] = ""
    saida: str
    observacoes: Optional[str] = ""


class FolhaCreate(BaseModel):
    funcionario_id: str
    mes: int
    ano: int
    salario_base: float
    horas_extras: float = 0
    valor_hora_extra: float = 0
    adicional_noturno: float = 0
    comissoes: float = 0
    vale_transporte: float = 0
    vale_alimentacao: float = 0
    plano_saude: float = 0
    outros_descontos: float = 0
    salario_bruto: float = 0
    inss: float = 0
    irpf: float = 0
    fgts: float = 0
    total_descontos: float = 0
    salario_liquido: float = 0


class FeriasCreate(BaseModel):
    funcionario_id: str
    data_inicio: str
    data_fim: str
    dias_vendidos: int = 0
    observacoes: Optional[str] = None


class EPIItem(BaseModel):
    nome: str
    ca: Optional[str] = None
    validade: Optional[str] = None
    data_entrega: Optional[str] = None


class FichaEPICreate(BaseModel):
    funcionario_id: str
    cargo: Optional[str] = None
    codigo_cbo: Optional[str] = None
    ocupacao_cbo: Optional[str] = None
    data_entrega: Optional[str] = None
    epis: List[EPIItem] = []
    observacoes: Optional[str] = None


# ===== CBO DATABASE =====
CBO_DATABASE = {
    "7152-10": {"titulo": "Pedreiro", "familia": "Trabalhadores de estruturas de alvenaria", "sinonimos": ["Pedreiro de alvenaria", "Pedreiro de edificações"]},
    "7152-15": {"titulo": "Pedreiro de acabamento", "familia": "Trabalhadores de estruturas de alvenaria", "sinonimos": ["Pedreiro de revestimento"]},
    "7152-20": {"titulo": "Pedreiro de fachadas", "familia": "Trabalhadores de estruturas de alvenaria", "sinonimos": []},
    "7241-05": {"titulo": "Eletricista de instalações (prediais)", "familia": "Eletricistas de instalações", "sinonimos": ["Eletricista predial", "Eletricista de manutenção predial"]},
    "7241-10": {"titulo": "Eletricista de instalações (industriais)", "familia": "Eletricistas de instalações", "sinonimos": ["Eletricista industrial"]},
    "7244-10": {"titulo": "Eletricista de manutenção de linhas elétricas", "familia": "Trabalhadores de manutenção de linhas elétricas", "sinonimos": []},
    "7153-05": {"titulo": "Armador de estrutura de concreto", "familia": "Trabalhadores de estruturas de concreto armado", "sinonimos": ["Armador de ferragens", "Ferreiro armador"]},
    "7153-10": {"titulo": "Moldador de concreto armado", "familia": "Trabalhadores de estruturas de concreto armado", "sinonimos": ["Carpinteiro de formas"]},
    "7155-05": {"titulo": "Carpinteiro", "familia": "Carpinteiros", "sinonimos": ["Carpinteiro de obras", "Carpinteiro de construção"]},
    "7155-35": {"titulo": "Carpinteiro de telhados", "familia": "Carpinteiros", "sinonimos": ["Carpinteiro de estruturas de madeira"]},
    "7161-10": {"titulo": "Pintor de obras", "familia": "Pintores de obras e decoradores", "sinonimos": ["Pintor de construção civil", "Pintor de paredes"]},
    "7161-20": {"titulo": "Pintor de estruturas metálicas", "familia": "Pintores de obras e decoradores", "sinonimos": []},
    "7166-10": {"titulo": "Encanador", "familia": "Instaladores de tubulações", "sinonimos": ["Bombeiro hidráulico", "Encanador industrial"]},
    "7166-20": {"titulo": "Instalador de tubulações de gás", "familia": "Instaladores de tubulações", "sinonimos": []},
    "7170-20": {"titulo": "Servente de obras", "familia": "Ajudantes de obras civis", "sinonimos": ["Auxiliar de pedreiro", "Ajudante de construção"]},
    "7170-25": {"titulo": "Vibradorista", "familia": "Ajudantes de obras civis", "sinonimos": []},
    "7823-10": {"titulo": "Operador de escavadeira", "familia": "Operadores de máquinas de movimentação de cargas", "sinonimos": ["Operador de retroescavadeira"]},
    "7823-20": {"titulo": "Operador de pá carregadeira", "familia": "Operadores de máquinas de movimentação de cargas", "sinonimos": []},
    "7824-10": {"titulo": "Operador de guindaste (móvel)", "familia": "Operadores de guindastes", "sinonimos": ["Guindasteiro"]},
    "7825-10": {"titulo": "Operador de empilhadeira", "familia": "Operadores de empilhadeiras", "sinonimos": ["Empilhadeirista"]},
    "7821-25": {"titulo": "Motorista de caminhão", "familia": "Condutores de veículos sobre rodas", "sinonimos": ["Caminhoneiro"]},
    "7821-15": {"titulo": "Motorista de ônibus", "familia": "Condutores de veículos sobre rodas", "sinonimos": []},
    "7251-10": {"titulo": "Soldador", "familia": "Soldadores e oxicortadores", "sinonimos": ["Soldador elétrico", "Soldador a arco"]},
    "7251-25": {"titulo": "Soldador a oxigás", "familia": "Soldadores e oxicortadores", "sinonimos": []},
    "7252-05": {"titulo": "Serralheiro", "familia": "Trabalhadores de estruturas de metal", "sinonimos": ["Serralheiro de alumínio", "Serralheiro industrial"]},
    "7211-10": {"titulo": "Ferramenteiro", "familia": "Ferramenteiros e operadores de máquinas", "sinonimos": ["Ferramenteiro de moldes"]},
    "7211-20": {"titulo": "Torneiro mecânico", "familia": "Ferramenteiros e operadores de máquinas", "sinonimos": ["Torneiro CNC"]},
    "7211-25": {"titulo": "Fresador mecânico", "familia": "Ferramenteiros e operadores de máquinas", "sinonimos": []},
    "5143-20": {"titulo": "Vigia", "familia": "Trabalhadores de segurança e vigilância", "sinonimos": ["Vigilante", "Guarda"]},
    "4110-10": {"titulo": "Auxiliar de escritório", "familia": "Escriturários em geral", "sinonimos": ["Auxiliar administrativo"]},
}

# EPIs por tipo de função
EPI_POR_CARGO = {
    "construcao_civil": [
        {"nome": "Capacete", "ca": "A definir", "validade_meses": 36, "prioridade": "Alta"},
        {"nome": "Óculos de proteção", "ca": "A definir", "validade_meses": 24, "prioridade": "Alta"},
        {"nome": "Luvas de proteção", "ca": "A definir", "validade_meses": 6, "prioridade": "Alta"},
        {"nome": "Botina de segurança", "ca": "A definir", "validade_meses": 12, "prioridade": "Alta"},
        {"nome": "Protetor auricular", "ca": "A definir", "validade_meses": 6, "prioridade": "Média"},
        {"nome": "Cinto de segurança", "ca": "A definir", "validade_meses": 24, "prioridade": "Alta"},
    ],
    "eletricidade": [
        {"nome": "Capacete classe B", "ca": "A definir", "validade_meses": 36, "prioridade": "Alta"},
        {"nome": "Óculos de proteção", "ca": "A definir", "validade_meses": 24, "prioridade": "Alta"},
        {"nome": "Luvas isolantes", "ca": "A definir", "validade_meses": 6, "prioridade": "Alta"},
        {"nome": "Botina isolante", "ca": "A definir", "validade_meses": 12, "prioridade": "Alta"},
        {"nome": "Vestimenta antichama", "ca": "A definir", "validade_meses": 24, "prioridade": "Alta"},
        {"nome": "Detector de tensão", "ca": "A definir", "validade_meses": 12, "prioridade": "Alta"},
    ],
    "soldagem": [
        {"nome": "Máscara de solda", "ca": "A definir", "validade_meses": 24, "prioridade": "Alta"},
        {"nome": "Avental de raspa", "ca": "A definir", "validade_meses": 12, "prioridade": "Alta"},
        {"nome": "Luvas de raspa", "ca": "A definir", "validade_meses": 3, "prioridade": "Alta"},
        {"nome": "Mangote de raspa", "ca": "A definir", "validade_meses": 6, "prioridade": "Alta"},
        {"nome": "Perneira de raspa", "ca": "A definir", "validade_meses": 12, "prioridade": "Alta"},
        {"nome": "Botina de segurança", "ca": "A definir", "validade_meses": 12, "prioridade": "Alta"},
    ],
    "operador_maquinas": [
        {"nome": "Capacete", "ca": "A definir", "validade_meses": 36, "prioridade": "Alta"},
        {"nome": "Protetor auricular", "ca": "A definir", "validade_meses": 6, "prioridade": "Alta"},
        {"nome": "Óculos de proteção", "ca": "A definir", "validade_meses": 24, "prioridade": "Média"},
        {"nome": "Luvas de vaqueta", "ca": "A definir", "validade_meses": 6, "prioridade": "Média"},
        {"nome": "Botina de segurança", "ca": "A definir", "validade_meses": 12, "prioridade": "Alta"},
    ],
    "administrativo": [
        {"nome": "Não requer EPIs específicos", "ca": "-", "validade_meses": 0, "prioridade": "Baixa"},
    ]
}


def get_categoria_epi(codigo_cbo: str) -> str:
    """Determina a categoria de EPI baseado no código CBO"""
    if codigo_cbo.startswith("715"):  # Construção civil
        return "construcao_civil"
    elif codigo_cbo.startswith("724"):  # Eletricidade
        return "eletricidade"
    elif codigo_cbo.startswith("725"):  # Serralheria e Metalurgia
        return "soldagem"
    elif codigo_cbo.startswith("782") or codigo_cbo.startswith("783"):  # Operadores
        return "operador_maquinas"
    elif codigo_cbo.startswith("411") or codigo_cbo.startswith("514"):  # Administrativo
        return "administrativo"
    else:
        return "construcao_civil"  # Default


# ===== DASHBOARD =====
@rh_router.get("/dashboard")
async def get_rh_dashboard():
    """Dashboard do RH com estatísticas"""
    try:
        total_funcionarios = await funcionarios_collection.count_documents({})
        funcionarios_ativos = await funcionarios_collection.count_documents({"status": "ativo"})
        funcionarios_ferias = await funcionarios_collection.count_documents({"status": "ferias"})
        funcionarios_afastados = await funcionarios_collection.count_documents({"status": "afastado"})
        
        total_folha = 0
        async for func in funcionarios_collection.find({"status": "ativo"}):
            total_folha += func.get("salario", 0)
        
        hoje = datetime.now()
        mes_atual = hoje.month
        
        aniversariantes = []
        async for func in funcionarios_collection.find({"status": "ativo"}):
            if func.get("data_nascimento"):
                try:
                    data_nasc = datetime.strptime(func["data_nascimento"], "%Y-%m-%d")
                    if data_nasc.month == mes_atual:
                        aniversariantes.append({
                            "nome": func["nome"],
                            "data": data_nasc.strftime("%d/%m"),
                            "cargo": func.get("cargo", "-")
                        })
                except:
                    pass
        
        alertas_ferias = []
        # Carrega dispensas: funcionários cujos alertas o admin marcou para ocultar
        dispensados_ids = set()
        async for d in db.ferias_alertas_dispensados.find({}, {"_id": 0, "funcionario_id": 1}):
            dispensados_ids.add(d.get("funcionario_id"))

        # Carrega última férias de cada funcionário — usado para evitar alertar quem já saiu de férias
        ultima_ferias_por_func = {}
        async for fe in db.ferias.find(
            {}, {"_id": 0, "funcionario_id": 1, "data_inicio": 1, "data_fim": 1}
        ).sort("data_inicio", -1):
            fid = fe.get("funcionario_id")
            if fid and fid not in ultima_ferias_por_func:
                ultima_ferias_por_func[fid] = fe.get("data_fim") or fe.get("data_inicio")

        async for func in funcionarios_collection.find({"status": "ativo"}):
            if func.get("id") in dispensados_ids:
                continue  # admin descartou esse alerta
            try:
                ultima = ultima_ferias_por_func.get(func.get("id"))
                if ultima:
                    # Conta a partir do fim das últimas férias
                    base_data = datetime.strptime(ultima[:10], "%Y-%m-%d")
                    meses_desde_ultima = (
                        (hoje.year - base_data.year) * 12 + (hoje.month - base_data.month)
                    )
                    if meses_desde_ultima >= 11:
                        alertas_ferias.append({
                            "funcionario_id": func.get("id"),
                            "nome": func["nome"],
                            "meses": meses_desde_ultima,
                            "mensagem": f"{meses_desde_ultima} meses desde a última férias",
                        })
                elif func.get("data_admissao"):
                    # Nunca teve férias registradas — usa admissão como base
                    data_adm = datetime.strptime(func["data_admissao"], "%Y-%m-%d")
                    meses_trabalhados = (
                        (hoje.year - data_adm.year) * 12 + (hoje.month - data_adm.month)
                    )
                    if meses_trabalhados >= 11:
                        alertas_ferias.append({
                            "funcionario_id": func.get("id"),
                            "nome": func["nome"],
                            "meses": meses_trabalhados,
                            "mensagem": f"Completou {meses_trabalhados} meses sem férias registradas",
                        })
            except Exception:
                pass
        
        alertas_epi = []
        async for ficha in epi_fichas_collection.find({}):
            func = await funcionarios_collection.find_one({"id": ficha.get("funcionario_id")})
            if func and ficha.get("epis"):
                for epi in ficha["epis"]:
                    if epi.get("validade"):
                        try:
                            validade = datetime.strptime(epi["validade"], "%Y-%m-%d")
                            dias_restantes = (validade - hoje).days
                            if 0 < dias_restantes <= 30:
                                alertas_epi.append({
                                    "funcionario": func["nome"],
                                    "epi": epi["nome"],
                                    "dias_restantes": dias_restantes
                                })
                        except:
                            pass
        
        # Calcular ponto de hoje
        hoje_str = hoje.strftime("%Y-%m-%d")
        presentes = 0
        atrasados = 0
        async for reg in ponto_collection.find({"data": hoje_str}):
            if reg.get("entrada"):
                presentes += 1
                try:
                    h, m = map(int, reg["entrada"].split(":"))
                    if h > 8 or (h == 8 and m > 15):
                        atrasados += 1
                except:
                    pass
        ausentes = funcionarios_ativos - presentes
        
        return {
            "total_funcionarios": total_funcionarios,
            "funcionarios_ativos": funcionarios_ativos,
            "funcionarios_ferias": funcionarios_ferias,
            "funcionarios_afastados": funcionarios_afastados,
            "total_folha": total_folha,
            "aniversariantes": aniversariantes[:5],
            "aniversariantes_mes": aniversariantes[:5],
            "alertas_ferias": alertas_ferias,
            "alertas_epi": alertas_epi[:10],
            "jornada": JORNADA_PADRAO,
            "tabela_inss": TABELA_INSS_2025,
            "tabela_irpf": TABELA_IRPF_2025,
            "ponto_hoje": {
                "presentes": presentes,
                "ausentes": max(0, ausentes),
                "atrasados": atrasados
            }
        }
    except Exception as e:
        return {
            "total_funcionarios": 0,
            "funcionarios_ativos": 0,
            "funcionarios_ferias": 0,
            "funcionarios_afastados": 0,
            "total_folha": 0,
            "aniversariantes": [],
            "aniversariantes_mes": [],
            "alertas_ferias": [],
            "alertas_epi": [],
            "jornada": JORNADA_PADRAO,
            "tabela_inss": TABELA_INSS_2025,
            "tabela_irpf": TABELA_IRPF_2025,
            "ponto_hoje": {
                "presentes": 0,
                "ausentes": 0,
                "atrasados": 0
            }
        }


# ===== ALERTAS DE FÉRIAS — DISPENSA =====
@rh_router.post("/ferias/alertas/dispensar/{funcionario_id}")
async def dispensar_alerta_ferias(funcionario_id: str):
    """Marca o alerta de férias deste funcionário como dispensado.
    O alerta não voltará no Dashboard até que o admin o reabilite ou o funcionário tenha
    novas férias registradas que mudem a base de cálculo."""
    func = await funcionarios_collection.find_one({"id": funcionario_id}, {"_id": 0})
    if not func:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    await db.ferias_alertas_dispensados.update_one(
        {"funcionario_id": funcionario_id},
        {"$set": {
            "funcionario_id": funcionario_id,
            "nome": func.get("nome"),
            "dispensado_em": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    return {"ok": True, "funcionario_id": funcionario_id, "nome": func.get("nome")}


@rh_router.delete("/ferias/alertas/dispensar/{funcionario_id}")
async def reativar_alerta_ferias(funcionario_id: str):
    """Reabilita os alertas de férias para um funcionário (remove a dispensa)."""
    res = await db.ferias_alertas_dispensados.delete_one({"funcionario_id": funcionario_id})
    return {"ok": True, "removidos": res.deleted_count}


@rh_router.post("/ferias/alertas/dispensar-todos")
async def dispensar_todos_alertas_ferias():
    """Dispensa em massa TODOS os alertas de férias atualmente ativos."""
    hoje = datetime.now()
    count = 0
    async for func in funcionarios_collection.find({"status": "ativo"}, {"_id": 0}):
        await db.ferias_alertas_dispensados.update_one(
            {"funcionario_id": func["id"]},
            {"$set": {
                "funcionario_id": func["id"],
                "nome": func.get("nome"),
                "dispensado_em": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=True,
        )
        count += 1
    return {"ok": True, "dispensados": count}


@rh_router.delete("/ferias/alertas/dispensar-todos")
async def reativar_todos_alertas_ferias():
    """Reabilita TODOS os alertas dispensados (limpa a coleção de dispensados)."""
    res = await db.ferias_alertas_dispensados.delete_many({})
    return {"ok": True, "reativados": res.deleted_count}


@rh_router.get("/ferias/alertas/dispensados")
async def listar_dispensados_ferias():
    """Lista os funcionários com alertas dispensados."""
    docs = []
    async for d in db.ferias_alertas_dispensados.find({}, {"_id": 0}):
        docs.append(d)
    return docs


# ===== FUNCIONARIOS =====
@rh_router.get("/funcionarios")
async def list_funcionarios():
    """Listar todos os funcionários"""
    funcionarios = []
    async for func in funcionarios_collection.find({}):
        func["_id"] = str(func["_id"])
        funcionarios.append(func)
    return funcionarios


@rh_router.get("/funcionarios/{funcionario_id}")
async def get_funcionario(funcionario_id: str):
    """Obter funcionário por ID"""
    func = await funcionarios_collection.find_one({"id": funcionario_id})
    if not func:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    func["_id"] = str(func["_id"])
    return func


@rh_router.post("/funcionarios")
async def create_funcionario(data: FuncionarioCreate):
    """Criar novo funcionário"""
    func_doc = data.dict()
    func_doc["id"] = str(uuid.uuid4())
    func_doc["created_at"] = datetime.now().isoformat()
    func_doc["anexos"] = []
    
    await funcionarios_collection.insert_one(func_doc)
    func_doc["_id"] = str(func_doc.get("_id", ""))
    return func_doc


@rh_router.put("/funcionarios/{funcionario_id}")
async def update_funcionario(funcionario_id: str, data: FuncionarioCreate):
    """Atualizar funcionário"""
    existing = await funcionarios_collection.find_one({"id": funcionario_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    
    update_data = data.dict()
    update_data["updated_at"] = datetime.now().isoformat()
    
    await funcionarios_collection.update_one(
        {"id": funcionario_id},
        {"$set": update_data}
    )
    
    return {"message": "Funcionário atualizado com sucesso"}


@rh_router.delete("/funcionarios/{funcionario_id}")
async def delete_funcionario(funcionario_id: str):
    """Excluir funcionário"""
    existing = await funcionarios_collection.find_one({"id": funcionario_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    
    if existing.get("anexos"):
        for anexo in existing["anexos"]:
            try:
                if os.path.exists(anexo.get("path", "")):
                    os.remove(anexo["path"])
            except:
                pass
    
    await funcionarios_collection.delete_one({"id": funcionario_id})
    return {"message": "Funcionário excluído com sucesso"}


# ===== PONTO =====
@rh_router.get("/ponto")
async def list_ponto(
    data: Optional[str] = None,
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    mes: Optional[int] = None,
    ano: Optional[int] = None,
    funcionario_id: Optional[str] = None,
):
    """Listar registros de ponto com filtros flexíveis.
    
    Aceita (em ordem de prioridade):
    - `data`: dia exato (YYYY-MM-DD)
    - `data_inicio` + `data_fim`: período inclusivo
    - `mes` + `ano`: mês inteiro
    - sem filtro de data: últimos 30 dias
    
    Retorna `{ registros, resumo }` com presentes / ausentes / atrasados / abonados / total_minutos.
    """
    query = {}
    
    # Resolve período
    if data:
        query["data"] = data
        periodo_ini = data
        periodo_fim = data
    elif data_inicio and data_fim:
        query["data"] = {"$gte": data_inicio, "$lte": data_fim}
        periodo_ini = data_inicio
        periodo_fim = data_fim
    elif mes and ano:
        ini = f"{ano}-{mes:02d}-01"
        if mes == 12:
            fim_excl = f"{ano + 1}-01-01"
        else:
            fim_excl = f"{ano}-{mes + 1:02d}-01"
        query["data"] = {"$gte": ini, "$lt": fim_excl}
        periodo_ini = ini
        periodo_fim = fim_excl
    else:
        from datetime import timedelta
        hoje = datetime.now()
        ini = (hoje - timedelta(days=30)).strftime("%Y-%m-%d")
        fim = hoje.strftime("%Y-%m-%d")
        query["data"] = {"$gte": ini, "$lte": fim}
        periodo_ini = ini
        periodo_fim = fim
    
    if funcionario_id:
        query["funcionario_id"] = funcionario_id
    
    # Carrega abonos do mesmo período para indicar dias abonados
    abonos_query = {}
    if isinstance(query.get("data"), dict):
        abonos_query["data"] = query["data"]
    elif isinstance(query.get("data"), str):
        abonos_query["data"] = query["data"]
    abonos_set = set()
    async for ab in ponto_abonos_collection.find(abonos_query, {"_id": 0, "funcionario_id": 1, "data": 1}):
        abonos_set.add((ab["funcionario_id"], ab["data"]))

    # Pré-carrega jornadas + mapa funcionario→jornada para RECALCULAR
    # `minutos_previstos` e `saldo_minutos` na hora. Isso garante que mudanças
    # de carga horária do funcionário sejam refletidas mesmo em registros
    # antigos importados antes da atribuição da jornada.
    jornada_padrao_doc = await _get_or_create_jornada_padrao()
    jornadas_map_listponto: dict = {jornada_padrao_doc["id"]: jornada_padrao_doc}
    async for _j in jornadas_collection.find({}, {"_id": 0}):
        jornadas_map_listponto[_j["id"]] = _j
    func_jornada_map_listponto: dict = {}
    async for _f in funcionarios_collection.find({}, {"_id": 0, "id": 1, "jornada_id": 1}):
        _fjid = _f.get("jornada_id")
        func_jornada_map_listponto[_f["id"]] = (
            jornadas_map_listponto.get(_fjid) if _fjid else jornada_padrao_doc
        ) or jornada_padrao_doc

    # Carrega registros (excluindo _id)
    registros = []
    cache_funcionarios = {}
    async for reg in ponto_collection.find(query, {"_id": 0}).sort([("data", -1), ("entrada", 1)]):
        fid = reg.get("funcionario_id", "")
        # Resolver nome
        if fid.startswith("NAO_CADASTRADO::"):
            reg["funcionario_nome"] = reg.get("funcionario_nome_planilha", "-")
            reg["funcionario_cadastrado"] = False
        else:
            if fid not in cache_funcionarios:
                f = await funcionarios_collection.find_one({"id": fid}, {"_id": 0, "nome": 1, "cargo": 1, "departamento": 1})
                cache_funcionarios[fid] = f
            f = cache_funcionarios.get(fid)
            reg["funcionario_nome"] = f.get("nome") if f else (reg.get("funcionario_nome_planilha") or "-")
            reg["funcionario_cargo"] = f.get("cargo") if f else None
            reg["funcionario_departamento"] = f.get("departamento") if f else None
            reg["funcionario_cadastrado"] = bool(f)
        # Marca abonos
        reg["abonado"] = (fid, reg.get("data")) in abonos_set

        # Recalcula minutos_previstos com a jornada ATUAL do funcionário
        if fid.startswith("NAO_CADASTRADO::"):
            jornada_func_reg = jornada_padrao_doc
        else:
            jornada_func_reg = func_jornada_map_listponto.get(fid, jornada_padrao_doc)
        prev_calc = _jornada_minutos_previstos(jornada_func_reg, reg.get("dia_semana", 6))
        reg["minutos_previstos"] = prev_calc
        reg["saldo_minutos"] = int(reg.get("minutos_trabalhados", 0) or 0) - prev_calc
        reg["jornada_id"] = jornada_func_reg.get("id") if jornada_func_reg else None
        reg["jornada_nome"] = jornada_func_reg.get("nome") if jornada_func_reg else None

        registros.append(reg)
    
    # Calcular resumo
    presentes = sum(1 for r in registros if (r.get("minutos_trabalhados") or 0) > 0)
    abonados = sum(1 for r in registros if r.get("abonado"))
    # Ausentes: registros sem batidas/incompletos em dias úteis (seg-sex)
    ausentes = sum(
        1 for r in registros
        if (r.get("minutos_trabalhados") or 0) == 0
        and not r.get("abonado")
        and r.get("dia_semana", 6) <= 4
    )
    # Atrasados: trabalharam mas tem saldo negativo (sem abono)
    atrasados = sum(
        1 for r in registros
        if (r.get("minutos_trabalhados") or 0) > 0
        and (r.get("saldo_minutos") or 0) < 0
        and not r.get("abonado")
    )
    total_minutos_trab = sum((r.get("minutos_trabalhados") or 0) for r in registros)
    total_minutos_prev = sum((r.get("minutos_previstos") or 0) for r in registros)
    
    # Funcionários únicos no período (com algum registro)
    funcionarios_no_periodo = len({r.get("funcionario_id") for r in registros})
    
    return {
        "registros": registros,
        "resumo": {
            "presentes": presentes,
            "ausentes": ausentes,
            "atrasados": atrasados,
            "abonados": abonados,
            "total_registros": len(registros),
            "total_funcionarios": funcionarios_no_periodo,
            "minutos_trabalhados": total_minutos_trab,
            "minutos_previstos": total_minutos_prev,
            "saldo_minutos": total_minutos_trab - total_minutos_prev,
            "periodo_inicio": periodo_ini,
            "periodo_fim": periodo_fim,
        },
    }


@rh_router.post("/ponto")
async def create_ponto(data: PontoCreate):
    """Criar registro de ponto"""
    existing = await ponto_collection.find_one({
        "funcionario_id": data.funcionario_id,
        "data": data.data
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="Já existe registro para este funcionário nesta data")
    
    ponto_doc = data.dict()
    ponto_doc["id"] = str(uuid.uuid4())
    ponto_doc["created_at"] = datetime.now().isoformat()
    
    await ponto_collection.insert_one(ponto_doc)
    ponto_doc["_id"] = str(ponto_doc.get("_id", ""))
    return ponto_doc


@rh_router.put("/ponto/{ponto_id}")
async def update_ponto(ponto_id: str, data: PontoCreate):
    """Atualizar registro de ponto"""
    existing = await ponto_collection.find_one({"id": ponto_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Registro não encontrado")
    
    update_data = data.dict()
    update_data["updated_at"] = datetime.now().isoformat()
    
    await ponto_collection.update_one({"id": ponto_id}, {"$set": update_data})
    return {"message": "Registro atualizado"}


@rh_router.delete("/ponto/{ponto_id}")
async def delete_ponto(ponto_id: str):
    """Excluir registro de ponto"""
    result = await ponto_collection.delete_one({"id": ponto_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Registro não encontrado")
    return {"message": "Registro excluído com sucesso"}


# ===== IMPORTAÇÃO DE PLANILHA DE PONTO (.xls) =====
def _normalizar_nome(nome: str) -> str:
    """Normaliza nome para comparação: trim + lowercase + remover múltiplos espaços."""
    if not nome:
        return ""
    return " ".join(nome.strip().lower().split())


def _tokens_nome(nome: str) -> list:
    """Quebra um nome em tokens limpos (sem pontuação, lowercase)."""
    norm = _normalizar_nome(nome)
    return [t.strip(".,;-_") for t in norm.split() if t.strip(".,;-_")]


def _match_funcionario_por_nome(nome_planilha: str, funcs_db: dict):
    """Encontra funcionário cadastrado para um nome vindo da planilha do relógio.
    
    Estratégia em camadas (do mais restritivo ao mais permissivo):
    1. Match exato após normalização (já no dict).
    2. Subsequência de tokens com prefixo: cada token da planilha deve ser igual ou
       prefixo de algum token do cadastro, preservando a ordem. Ex.:
       "GUSTAVO RODRIGUES" → "GUSTAVO HENRIQUE A RODRIGUES" ✓
       "LUIZ CARLOS M." (token "m") → "LUIZ CARLOS MOURA ..." (prefixo "m" → "moura") ✓
    
    Em caso de múltiplos matches, retorna o de maior pontuação (mais tokens cobertos)
    e desempate pelo nome mais curto (cadastro mais "próximo" do que a planilha trouxe).
    """
    if not nome_planilha:
        return None
    
    target_norm = _normalizar_nome(nome_planilha)
    
    # 1) Match exato
    if target_norm in funcs_db:
        return funcs_db[target_norm]
    
    tokens_p = _tokens_nome(nome_planilha)
    if not tokens_p:
        return None
    
    candidatos = []
    for nome_norm, func in funcs_db.items():
        tokens_c = _tokens_nome(nome_norm)
        if not tokens_c:
            continue
        # Verifica subsequência com prefixo
        i = 0
        for tc in tokens_c:
            if i < len(tokens_p) and tc.startswith(tokens_p[i]):
                i += 1
        if i == len(tokens_p):
            # Pontuação: nº de tokens da planilha que casaram (=len(tokens_p) sempre que entra aqui)
            # Desempate: cadastro com MENOS tokens extras é mais provável de ser o correto
            score = len(tokens_p) * 1000 - len(tokens_c)
            candidatos.append((score, func))
    
    if not candidatos:
        return None
    candidatos.sort(key=lambda x: x[0], reverse=True)
    return candidatos[0][1]


def _parse_batidas_celula(valor: str) -> list:
    """Recebe o valor da célula (ex: '08:01\\n11:50\\n13:55\\n17:48') e retorna lista ordenada de batidas HH:MM."""
    if not valor:
        return []
    raw = str(valor).replace("\r", "\n").split("\n")
    batidas = []
    for b in raw:
        b = b.strip()
        if not b:
            continue
        # Aceita formatos HH:MM ou HH:MM:SS
        partes = b.split(":")
        if len(partes) >= 2:
            try:
                h = int(partes[0])
                m = int(partes[1])
                if 0 <= h <= 23 and 0 <= m <= 59:
                    batidas.append(f"{h:02d}:{m:02d}")
            except (ValueError, TypeError):
                continue
    # Ordenar
    batidas.sort()
    return batidas


def _calcular_minutos_trabalhados(batidas: list, dia_semana: int) -> tuple:
    """Calcula minutos trabalhados a partir de uma lista de batidas ordenadas.
    Regras:
      - 0 batidas: dia sem registro (falta se for dia útil)
      - 1 batida: incompleto (entrada apenas) → 0 minutos
      - 2 batidas: entrada e saída sem almoço (jornada curta, ex: sábado)
      - 3 batidas: assume entrada/saída-almoço/retorno (sem saída final) ou inverso → aplica almoço padrão de 60min
      - 4+ batidas: pareia em entrada/saída-almoço/retorno-almoço/saída
    Retorna: (minutos_trabalhados, status_legivel)
    """
    if not batidas:
        return 0, "sem_registro"
    if len(batidas) == 1:
        return 0, "incompleto"
    
    def to_min(s):
        h, m = s.split(":")
        return int(h) * 60 + int(m)
    
    pts = [to_min(b) for b in batidas]
    
    if len(pts) == 2:
        return max(0, pts[1] - pts[0]), "ok_2_batidas"
    
    if len(pts) == 3:
        # Inferir: entrada=p1, saída=p3; descontar 60min de almoço
        bruto = pts[2] - pts[0]
        liquido = max(0, bruto - 60)
        return liquido, "inferido_3_batidas"
    
    # 4+ batidas: usa as primeiras 4 (entrada, saída-almoço, retorno-almoço, saída)
    e, sa, ra, s = pts[0], pts[1], pts[2], pts[3]
    parte1 = max(0, sa - e)
    parte2 = max(0, s - ra)
    return parte1 + parte2, "ok_4_batidas"


def _jornada_prevista_minutos(dia_semana: int) -> int:
    """0=Seg ... 5=Sáb, 6=Dom. Retorna minutos previstos."""
    if dia_semana <= 4:  # Seg-Sex
        return 8 * 60
    if dia_semana == 5:  # Sábado
        return 4 * 60
    return 0  # Domingo


@rh_router.post("/ponto/importar-planilha")
async def importar_planilha_ponto(file: UploadFile = File(...)):
    """Importa planilha .xls/.xlsx de registro de presença (formato Topdata/Hikvision/etc.).
    
    Estrutura esperada:
      Linha N+0: 'IDUsuário: <id>' ... 'Nome: <nome>' ... 'Dep.: <departamento>'
      Linha N+1: cabeçalho com dias do mês (1, 2, 3, ..., 31)
      Linha N+2: batidas separadas por \\n para cada dia
      Linha 2 contém 'Data de presença:DD/MM/AAAA~DD/MM/AAAA'
    
    Comportamento:
      - Match por nome exato (case-insensitive normalizado)
      - Funcionários não encontrados: registros são salvos com flag 'funcionario_nao_cadastrado'
      - Sobrescreve APENAS os dias do intervalo da planilha (ex.: 01-15/04 não apaga 16-30/04)
      - Re-upload do mesmo intervalo substitui os registros existentes desses dias
    """
    import xlrd
    import re
    
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Arquivo vazio")
    
    # Detectar formato pelo magic byte
    if content[:8].hex().lower() == "d0cf11e0a1b11ae1":
        try:
            wb = xlrd.open_workbook(file_contents=content, formatting_info=False)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao abrir .xls: {e}")
    elif content[:4] == b"PK\x03\x04":
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            wb_x = load_workbook(BytesIO(content), data_only=True)
            # Converter para um wb-like com sheet_by_index
            class _SheetWrap:
                def __init__(self, ws):
                    self.ws = ws
                    self.nrows = ws.max_row
                    self.ncols = ws.max_column
                def cell_value(self, r, c):
                    v = self.ws.cell(row=r+1, column=c+1).value
                    return v if v is not None else ""
            class _WBWrap:
                def __init__(self, wbx):
                    self.wbx = wbx
                def sheet_by_index(self, i):
                    return _SheetWrap(self.wbx.worksheets[i])
            wb = _WBWrap(wb_x)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao abrir .xlsx: {e}")
    else:
        raise HTTPException(status_code=400, detail="Formato não reconhecido. Envie um arquivo .xls ou .xlsx válido.")
    
    sh = wb.sheet_by_index(0)
    
    # Detectar período (procurar nas primeiras 10 linhas)
    periodo_match = None
    for r in range(min(10, sh.nrows)):
        for c in range(sh.ncols):
            v = str(sh.cell_value(r, c) or "")
            m = re.search(r"(\d{2}/\d{2}/\d{4})\s*[~\-–]\s*(\d{2}/\d{2}/\d{4})", v)
            if m:
                periodo_match = m
                break
        if periodo_match:
            break
    
    if not periodo_match:
        raise HTTPException(status_code=400, detail="Não foi possível identificar o período (procurando 'Data de presença:DD/MM/AAAA~DD/MM/AAAA')")
    
    inicio_str, fim_str = periodo_match.group(1), periodo_match.group(2)
    try:
        inicio_dt = datetime.strptime(inicio_str, "%d/%m/%Y")
        fim_dt = datetime.strptime(fim_str, "%d/%m/%Y")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Período inválido: {e}")
    
    if inicio_dt.year != fim_dt.year or inicio_dt.month != fim_dt.month:
        raise HTTPException(status_code=400, detail="A planilha deve cobrir UM ÚNICO mês (limitação atual).")
    
    ano = inicio_dt.year
    mes = inicio_dt.month
    
    # Carregar todos os funcionários para matching por nome
    funcs_db = {}
    async for f in funcionarios_collection.find({}):
        funcs_db[_normalizar_nome(f.get("nome", ""))] = f
    
    # Limpeza por intervalo da planilha (e NÃO o mês inteiro).
    # Permite uploads parciais (ex.: 01-15/04) sem destruir dias importados antes
    # em outras planilhas (ex.: 16-30/04). Re-upload do mesmo intervalo sobrescreve.
    inicio_periodo_iso = inicio_dt.strftime("%Y-%m-%d")
    fim_periodo_iso = fim_dt.strftime("%Y-%m-%d")
    registros_sobrescritos = await ponto_collection.count_documents({
        "origem": "planilha_xls",
        "data": {"$gte": inicio_periodo_iso, "$lte": fim_periodo_iso},
    })
    await ponto_collection.delete_many({
        "origem": "planilha_xls",
        "data": {"$gte": inicio_periodo_iso, "$lte": fim_periodo_iso},
    })

    # Pré-carrega jornadas (Padrão + customizadas) para resolver a carga horária
    # correta de cada funcionário no momento da importação. Sem isso, todos os
    # registros são salvos com a jornada padrão hardcoded (Seg-Sex 8h, Sáb 4h).
    jornada_padrao_doc = await _get_or_create_jornada_padrao()
    jornadas_map: dict = {jornada_padrao_doc["id"]: jornada_padrao_doc}
    async for _j in jornadas_collection.find({}, {"_id": 0}):
        jornadas_map[_j["id"]] = _j

    # Iterar linhas em busca de blocos de funcionários
    funcionarios_processados = []
    funcionarios_nao_cadastrados = []
    total_registros_inseridos = 0
    
    r = 0
    while r < sh.nrows:
        linha_idu = [str(sh.cell_value(r, c) or "").strip() for c in range(sh.ncols)]
        linha_concat = " ".join(linha_idu).lower()
        
        # Buscar a linha que contém 'idusuário' ou 'idusuario'
        if "idusuário" in linha_concat or "idusuario" in linha_concat:
            # Extrair IDUsuário, Nome, Dep.
            id_usuario = ""
            nome = ""
            dep = ""
            for c in range(sh.ncols):
                v = str(sh.cell_value(r, c) or "").strip()
                v_low = v.lower()
                if v_low.startswith("idusuário") or v_low.startswith("idusuario"):
                    # ID está no próximo valor não vazio
                    for c2 in range(c + 1, sh.ncols):
                        nv = str(sh.cell_value(r, c2) or "").strip()
                        if nv:
                            id_usuario = nv
                            break
                elif v_low.startswith("nome"):
                    for c2 in range(c + 1, sh.ncols):
                        nv = str(sh.cell_value(r, c2) or "").strip()
                        if nv:
                            nome = nv
                            break
                elif v_low.startswith("dep"):
                    for c2 in range(c + 1, sh.ncols):
                        nv = str(sh.cell_value(r, c2) or "").strip()
                        if nv:
                            dep = nv
                            break
            
            # Próxima linha: cabeçalho de dias (1.0, 2.0, ..., 31.0)
            # Linha r+1 = dias, Linha r+2 = batidas
            if r + 2 < sh.nrows and nome:
                linha_dias = [sh.cell_value(r + 1, c) for c in range(sh.ncols)]
                linha_batidas = [sh.cell_value(r + 2, c) for c in range(sh.ncols)]
                
                # Match funcionário (exato → subsequência de tokens com prefixo)
                func = _match_funcionario_por_nome(nome, funcs_db)
                func_id = func["id"] if func else None
                cargo = func.get("cargo", "") if func else ""

                # Resolve a jornada deste funcionário (ou usa Padrão p/ não-cadastrado
                # / sem jornada_id definido). Isso garante que `minutos_previstos`
                # respeite a carga horária atribuída a ele.
                fjid = func.get("jornada_id") if func else None
                jornada_func = jornadas_map.get(fjid) if fjid else None
                if not jornada_func:
                    jornada_func = jornada_padrao_doc

                if not func:
                    funcionarios_nao_cadastrados.append({
                        "nome": nome,
                        "id_usuario_planilha": id_usuario,
                        "departamento_planilha": dep,
                    })
                
                # Iterar dias: cada coluna onde linha_dias é numérica = dia válido
                dias_processados = 0
                for c in range(sh.ncols):
                    dia_val = linha_dias[c]
                    try:
                        dia_num = int(float(dia_val)) if dia_val not in (None, "") else 0
                    except (ValueError, TypeError):
                        dia_num = 0
                    
                    if not (1 <= dia_num <= 31):
                        continue
                    
                    # Validar que o dia existe no mês
                    try:
                        data_str = f"{ano}-{mes:02d}-{dia_num:02d}"
                        data_dt = datetime.strptime(data_str, "%Y-%m-%d")
                    except ValueError:
                        continue

                    # Considerar apenas dias dentro do intervalo declarado pela planilha
                    if data_dt < inicio_dt or data_dt > fim_dt:
                        continue
                    
                    batidas_raw = linha_batidas[c] if c < len(linha_batidas) else ""
                    batidas = _parse_batidas_celula(batidas_raw)
                    
                    dia_semana = data_dt.weekday()
                    minutos_trab, status_dia = _calcular_minutos_trabalhados(batidas, dia_semana)
                    minutos_prev = _jornada_minutos_previstos(jornada_func, dia_semana)
                    saldo = minutos_trab - minutos_prev
                    
                    # Montar registro: usa primeira batida como entrada, última como saída, etc.
                    entrada = batidas[0] if len(batidas) >= 1 else ""
                    saida = batidas[-1] if len(batidas) >= 2 else ""
                    saida_almoco = batidas[1] if len(batidas) >= 4 else (batidas[1] if len(batidas) == 3 else "")
                    retorno_almoco = batidas[2] if len(batidas) >= 4 else ""
                    
                    ponto_doc = {
                        "id": str(uuid.uuid4()),
                        "funcionario_id": func_id or f"NAO_CADASTRADO::{_normalizar_nome(nome)}",
                        "funcionario_nome_planilha": nome,
                        "funcionario_nao_cadastrado": func is None,
                        "id_usuario_planilha": id_usuario,
                        "departamento_planilha": dep,
                        "data": data_str,
                        "batidas": batidas,
                        "entrada": entrada,
                        "saida_almoco": saida_almoco,
                        "retorno_almoco": retorno_almoco,
                        "saida": saida,
                        "minutos_trabalhados": minutos_trab,
                        "minutos_previstos": minutos_prev,
                        "saldo_minutos": saldo,
                        "status_dia": status_dia,
                        "dia_semana": dia_semana,
                        "jornada_id": jornada_func.get("id"),
                        "jornada_nome": jornada_func.get("nome"),
                        "origem": "planilha_xls",
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    }
                    
                    # Sobrescrever existente
                    await ponto_collection.delete_many({
                        "funcionario_id": ponto_doc["funcionario_id"],
                        "data": data_str,
                    })
                    await ponto_collection.insert_one(ponto_doc)
                    total_registros_inseridos += 1
                    dias_processados += 1
                
                funcionarios_processados.append({
                    "nome": nome,
                    "id_usuario_planilha": id_usuario,
                    "departamento": dep or cargo,
                    "funcionario_id": func_id,
                    "cadastrado": func is not None,
                    "dias_com_registro": dias_processados,
                })
                r += 3
                continue
        r += 1
    
    msg_base = f"Importação concluída: {total_registros_inseridos} registros de ponto importados."
    if registros_sobrescritos > 0:
        msg_base += f" {registros_sobrescritos} registro(s) anterior(es) do mesmo período foram sobrescritos."

    return {
        "message": msg_base,
        "periodo": {"inicio": inicio_str, "fim": fim_str, "mes": mes, "ano": ano},
        "total_funcionarios": len(funcionarios_processados),
        "total_registros": total_registros_inseridos,
        "registros_sobrescritos": registros_sobrescritos,
        "funcionarios_processados": funcionarios_processados,
        "funcionarios_nao_cadastrados": funcionarios_nao_cadastrados,
        "aviso_nao_cadastrados": (
            f"{len(funcionarios_nao_cadastrados)} funcionário(s) da planilha não estão cadastrados na plataforma "
            f"(serão exibidos no quadro com aviso): {', '.join([f['nome'] for f in funcionarios_nao_cadastrados[:5]])}"
            if funcionarios_nao_cadastrados else None
        ),
    }


@rh_router.get("/ponto/dashboard-mensal")
async def get_ponto_dashboard_mensal(mes: int, ano: int):
    """Quadro consolidado por funcionário do mês: horas trabalhadas vs previstas, saldo, banco de horas acumulado.
    Inclui detalhamento dia a dia."""
    inicio_mes = f"{ano}-{mes:02d}-01"
    if mes == 12:
        fim_mes = f"{ano + 1}-01-01"
    else:
        fim_mes = f"{ano}-{mes + 1:02d}-01"
    
    pipeline = [
        {"$match": {"data": {"$gte": inicio_mes, "$lt": fim_mes}}},
        {"$sort": {"data": 1}},
    ]
    
    por_func = {}
    async for reg in ponto_collection.aggregate(pipeline):
        fid = reg["funcionario_id"]
        if fid not in por_func:
            por_func[fid] = []
        por_func[fid].append(reg)
    
    # Banco de horas acumulado: vamos calcular usando a jornada CORRENTE de cada funcionário
    # (aplicação retroativa). Para isso, precisamos primeiro carregar jornadas e funcionários.
    jornada_padrao = await _get_or_create_jornada_padrao()
    jornadas_map = {jornada_padrao["id"]: jornada_padrao}
    async for j in jornadas_collection.find({}, {"_id": 0}):
        jornadas_map[j["id"]] = j
    
    func_jornada_map = {}  # funcionario_id -> jornada_doc
    async for f in funcionarios_collection.find({}, {"_id": 0, "id": 1, "jornada_id": 1}):
        fjid = f.get("jornada_id")
        func_jornada_map[f["id"]] = jornadas_map.get(fjid) if fjid else jornada_padrao
    
    banco_acumulado_por_func = {}
    async for r in ponto_collection.find(
        {"data": {"$lt": fim_mes}},
        {"_id": 0, "funcionario_id": 1, "data": 1, "minutos_trabalhados": 1, "dia_semana": 1},
    ):
        fid = r["funcionario_id"]
        j_func = func_jornada_map.get(fid, jornada_padrao)
        prev = _jornada_minutos_previstos(j_func, r.get("dia_semana", 6))
        trab = int(r.get("minutos_trabalhados", 0) or 0)
        banco_acumulado_por_func[fid] = banco_acumulado_por_func.get(fid, 0) + (trab - prev)
    
    # Carregar TODOS os abonos (não só do mês) para neutralizar saldos no banco acumulado
    abonos_global_por_func_data = {}
    async for ab in ponto_abonos_collection.find({}, {"_id": 0}):
        abonos_global_por_func_data[(ab["funcionario_id"], ab["data"])] = ab
    
    # Para cada (funcionario, dia abonado em qualquer época), descontar do banco acumulado o saldo
    # daquele dia (que provavelmente é negativo) para neutralizá-lo.
    if abonos_global_por_func_data:
        async for r in ponto_collection.find(
            {"data": {"$lt": fim_mes}}, {"_id": 0, "funcionario_id": 1, "data": 1, "minutos_trabalhados": 1, "dia_semana": 1}
        ):
            chave = (r["funcionario_id"], r["data"])
            if chave in abonos_global_por_func_data:
                fid = r["funcionario_id"]
                j_func = func_jornada_map.get(fid, jornada_padrao)
                prev = _jornada_minutos_previstos(j_func, r.get("dia_semana", 6))
                trab = int(r.get("minutos_trabalhados", 0) or 0)
                saldo_dia = trab - prev
                banco_acumulado_por_func[fid] = banco_acumulado_por_func.get(fid, 0) - saldo_dia
    
    # Carregar abonos do mês (chave: (funcionario_id, data))
    abonos_por_func_data = {
        k: v for k, v in abonos_global_por_func_data.items()
        if inicio_mes <= k[1] < fim_mes
    }
    
    funcionarios_dashboard = []
    for fid, registros in por_func.items():
        # Tentar achar funcionário
        nome = ""
        cargo = ""
        departamento = ""
        cadastrado = False
        jornada_func = None
        if fid.startswith("NAO_CADASTRADO::"):
            nome = registros[0].get("funcionario_nome_planilha", "")
            departamento = registros[0].get("departamento_planilha", "")
            cargo = "(Não cadastrado)"
            cadastrado = False
            jornada_func = jornada_padrao  # Não cadastrados usam padrão
        else:
            func = await funcionarios_collection.find_one({"id": fid})
            if func:
                nome = func.get("nome", "")
                cargo = func.get("cargo", "")
                departamento = func.get("departamento", "") or registros[0].get("departamento_planilha", "")
                cadastrado = True
                # Resolve jornada: se funcionário tem jornada_id válida, usa; senão padrão
                fjid = func.get("jornada_id")
                jornada_func = jornadas_map.get(fjid) if fjid else None
                if not jornada_func:
                    jornada_func = jornada_padrao
            else:
                nome = registros[0].get("funcionario_nome_planilha", "?")
                cargo = "(Funcionário removido)"
                cadastrado = False
                jornada_func = jornada_padrao
        
        # Garante consistência: usa também o map global (já calculado para o banco acumulado)
        if cadastrado and not jornada_func:
            jornada_func = func_jornada_map.get(fid, jornada_padrao)
        
        # Aplicar abonos: dias abonados têm saldo neutralizado para 0 (trabalhadas tratadas como previstas)
        # Recalcular minutos_previstos via jornada do funcionário (substitui o calculado na importação)
        total_trab_real = sum(int(r.get("minutos_trabalhados", 0) or 0) for r in registros)
        total_prev = sum(
            _jornada_minutos_previstos(jornada_func, r.get("dia_semana", 6))
            for r in registros
        )
        
        # Soma saldos dia a dia, neutralizando dias abonados
        saldo_mes = 0
        dias_abonados = 0
        for r in registros:
            data_r = r.get("data")
            ab = abonos_por_func_data.get((fid, data_r))
            if ab:
                # Dia abonado: contribuição zero para o saldo
                dias_abonados += 1
            else:
                trab_dia = int(r.get("minutos_trabalhados", 0) or 0)
                prev_dia = _jornada_minutos_previstos(jornada_func, r.get("dia_semana", 6))
                saldo_mes += (trab_dia - prev_dia)
        
        # Dias incompletos / faltas considerando jornada
        dias_com_registro = sum(1 for r in registros if r.get("minutos_trabalhados", 0) > 0)
        dias_incompletos = sum(
            1 for r in registros
            if r.get("status_dia") == "incompleto" and (fid, r.get("data")) not in abonos_por_func_data
        )
        dias_falta = sum(
            1 for r in registros
            if r.get("status_dia") == "sem_registro"
            and _jornada_minutos_previstos(jornada_func, r.get("dia_semana", 6)) > 0
            and (fid, r.get("data")) not in abonos_por_func_data
        )
        
        # Detalhamento dia a dia (com info de abono)
        detalhe = []
        for r in registros:
            data_r = r.get("data")
            ab = abonos_por_func_data.get((fid, data_r))
            trab_dia = int(r.get("minutos_trabalhados", 0) or 0)
            prev_dia = _jornada_minutos_previstos(jornada_func, r.get("dia_semana", 6))
            saldo_dia = trab_dia - prev_dia
            if ab:
                saldo_dia = 0  # Abono neutraliza
            detalhe.append({
                "data": data_r,
                "dia_semana": r.get("dia_semana"),
                "batidas": r.get("batidas", []),
                "minutos_trabalhados": trab_dia,
                "minutos_previstos": prev_dia,
                "saldo_minutos": saldo_dia,
                "status_dia": "abonado" if ab else r.get("status_dia"),
                "abono": (
                    {"tipo": ab["tipo"], "motivo": ab["motivo"], "id": ab["id"], "anexo": ab.get("anexo")}
                    if ab else None
                ),
            })
        
        # total_trab "efetivo" considera abonos como horas previstas cumpridas (visualmente fecha a barra)
        minutos_abonados_compensados = sum(
            _jornada_minutos_previstos(jornada_func, r.get("dia_semana", 6)) - int(r.get("minutos_trabalhados", 0) or 0)
            for r in registros
            if (fid, r.get("data")) in abonos_por_func_data
        )
        total_trab = total_trab_real + max(0, minutos_abonados_compensados)
        
        # Banco acumulado: já vem com os abonos neutralizados de todos os meses
        banco_total = banco_acumulado_por_func.get(fid, 0)
        
        # Carrega observação do mês para este funcionário
        obs_doc = await ponto_observacoes_collection.find_one(
            {"funcionario_id": fid, "mes": mes, "ano": ano}, {"_id": 0}
        )
        observacao_texto = obs_doc.get("texto", "") if obs_doc else ""
        
        # Lista de abonos do mês desse funcionário
        abonos_func = [
            {**ab}
            for k, ab in abonos_por_func_data.items()
            if k[0] == fid
        ]
        
        funcionarios_dashboard.append({
            "funcionario_id": fid,
            "nome": nome,
            "cargo": cargo,
            "departamento": departamento,
            "cadastrado": cadastrado,
            "jornada_id": jornada_func.get("id") if jornada_func else None,
            "jornada_nome": jornada_func.get("nome") if jornada_func else "Padrão",
            "minutos_trabalhados": total_trab,
            "minutos_previstos": total_prev,
            "saldo_mes_minutos": saldo_mes,
            "banco_horas_acumulado_minutos": banco_total,
            "dias_com_registro": dias_com_registro,
            "dias_incompletos": dias_incompletos,
            "dias_falta": dias_falta,
            "dias_abonados": dias_abonados,
            "detalhe_dias": detalhe,
            "observacao": observacao_texto,
            "abonos": abonos_func,
        })
    
    # Ordenar: cadastrados primeiro, por nome
    funcionarios_dashboard.sort(key=lambda x: (not x["cadastrado"], x["nome"].lower()))
    
    # Totais gerais
    total_funcionarios = len(funcionarios_dashboard)
    total_extras = sum(max(0, f["saldo_mes_minutos"]) for f in funcionarios_dashboard)
    total_devidas = sum(abs(min(0, f["saldo_mes_minutos"])) for f in funcionarios_dashboard)
    
    return {
        "mes": mes,
        "ano": ano,
        "total_funcionarios": total_funcionarios,
        "total_extras_minutos": total_extras,
        "total_devidas_minutos": total_devidas,
        "funcionarios": funcionarios_dashboard,
    }


# ============================================================
# BANCO DE HORAS - Visão dedicada baseada nos dados do Ponto
# ============================================================
async def _calcular_banco_horas_por_funcionario(
    ate_data: Optional[str] = None,
    de_data: Optional[str] = None,
) -> dict:
    """Calcula o saldo do banco de horas de cada funcionário com base nos
    registros de ponto entre `de_data` e `ate_data` (ambos inclusive).
    Sempre soma os AJUSTES MANUAIS (banco_horas_ajustes) feitos no período.
    Retorna dict {funcionario_id: {"saldo_minutos", "primeiro_registro", "ultimo_registro", "dias_registrados", "ajustes_minutos"}}.
    """
    if ate_data is None:
        ate_data = datetime.now().strftime("%Y-%m-%d")
    fim_str = ate_data + "T99"  # garante incluir o dia atual

    # Filtro de data sobre `data` (string YYYY-MM-DD funciona com $lte/$gte)
    data_filter: dict = {"$lte": ate_data}
    if de_data:
        data_filter["$gte"] = de_data

    # Carrega jornadas
    jornada_padrao = await _get_or_create_jornada_padrao()
    jornadas_map = {jornada_padrao["id"]: jornada_padrao}
    async for j in jornadas_collection.find({}, {"_id": 0}):
        jornadas_map[j["id"]] = j

    func_jornada_map = {}
    async for f in funcionarios_collection.find({}, {"_id": 0, "id": 1, "jornada_id": 1}):
        fjid = f.get("jornada_id")
        func_jornada_map[f["id"]] = jornadas_map.get(fjid) if fjid else jornada_padrao

    # Acumular saldo dia a dia
    resumo: dict = {}  # fid -> {saldo, primeiro, ultimo, dias, ajustes}
    async for r in ponto_collection.find(
        {"data": data_filter},
        {"_id": 0, "funcionario_id": 1, "data": 1, "minutos_trabalhados": 1, "dia_semana": 1},
    ):
        fid = r["funcionario_id"]
        j_func = func_jornada_map.get(fid, jornada_padrao)
        prev = _jornada_minutos_previstos(j_func, r.get("dia_semana", 6))
        trab = int(r.get("minutos_trabalhados", 0) or 0)
        saldo_dia = trab - prev
        if fid not in resumo:
            resumo[fid] = {
                "saldo_minutos": 0,
                "primeiro_registro": r["data"],
                "ultimo_registro": r["data"],
                "dias_registrados": 0,
                "ajustes_minutos": 0,
            }
        resumo[fid]["saldo_minutos"] += saldo_dia
        if r["data"] < resumo[fid]["primeiro_registro"]:
            resumo[fid]["primeiro_registro"] = r["data"]
        if r["data"] > resumo[fid]["ultimo_registro"]:
            resumo[fid]["ultimo_registro"] = r["data"]
        resumo[fid]["dias_registrados"] += 1

    # Neutraliza saldos de dias com abono
    abonos_filter: dict = {"data": {"$lte": ate_data}}
    if de_data:
        abonos_filter["data"]["$gte"] = de_data
    async for ab in ponto_abonos_collection.find(abonos_filter, {"_id": 0}):
        fid = ab["funcionario_id"]
        if fid not in resumo:
            continue
        ponto_dia = await ponto_collection.find_one(
            {"funcionario_id": fid, "data": ab["data"]}, {"_id": 0}
        )
        if not ponto_dia:
            continue
        j_func = func_jornada_map.get(fid, jornada_padrao)
        prev = _jornada_minutos_previstos(j_func, ponto_dia.get("dia_semana", 6))
        trab = int(ponto_dia.get("minutos_trabalhados", 0) or 0)
        saldo_dia = trab - prev
        resumo[fid]["saldo_minutos"] -= saldo_dia

    # Aplica ajustes manuais do banco de horas
    ajustes_filter: dict = {"data": {"$lte": ate_data}}
    if de_data:
        ajustes_filter["data"]["$gte"] = de_data
    async for aj in db.banco_horas_ajustes.find(ajustes_filter, {"_id": 0}):
        fid = aj.get("funcionario_id")
        if not fid:
            continue
        if fid not in resumo:
            resumo[fid] = {
                "saldo_minutos": 0, "primeiro_registro": aj.get("data"),
                "ultimo_registro": aj.get("data"), "dias_registrados": 0,
                "ajustes_minutos": 0,
            }
        delta = int(aj.get("minutos") or 0)
        resumo[fid]["saldo_minutos"] += delta
        resumo[fid]["ajustes_minutos"] += delta

    return resumo


@rh_router.get("/banco-horas/resumo")
async def banco_horas_resumo(
    ate_data: Optional[str] = None,
    de_data: Optional[str] = None,
):
    """Lista todos os funcionários com saldo acumulado de banco de horas no
    período informado. `de_data` opcional para apurar apenas um intervalo."""
    saldos = await _calcular_banco_horas_por_funcionario(ate_data=ate_data, de_data=de_data)
    
    funcionarios = []
    async for f in funcionarios_collection.find({}, {"_id": 0}):
        s = saldos.get(f["id"], {})
        funcionarios.append({
            "funcionario_id": f["id"],
            "nome": f.get("nome", "-"),
            "cargo": f.get("cargo", "-"),
            "departamento": f.get("departamento", "-"),
            "status": f.get("status", "ativo"),
            "data_admissao": f.get("data_admissao"),
            "saldo_minutos": s.get("saldo_minutos", 0),
            "ajustes_minutos": s.get("ajustes_minutos", 0),
            "dias_registrados": s.get("dias_registrados", 0),
            "primeiro_registro": s.get("primeiro_registro"),
            "ultimo_registro": s.get("ultimo_registro"),
        })
    funcionarios.sort(key=lambda x: x["nome"].lower())
    
    total_credito = sum(max(0, f["saldo_minutos"]) for f in funcionarios)
    total_debito = sum(abs(min(0, f["saldo_minutos"])) for f in funcionarios)
    
    return {
        "de_data": de_data,
        "ate_data": ate_data or datetime.now().strftime("%Y-%m-%d"),
        "total_funcionarios": len(funcionarios),
        "total_credito_minutos": total_credito,
        "total_debito_minutos": total_debito,
        "saldo_liquido_minutos": total_credito - total_debito,
        "funcionarios": funcionarios,
    }


# ============ AJUSTES MANUAIS DE BANCO DE HORAS ============
class BancoHorasAjusteCreate(BaseModel):
    funcionario_id: str
    minutos: int  # positivo: adiciona; negativo: retira
    data: Optional[str] = None  # YYYY-MM-DD; default = hoje
    motivo: str
    tipo: Optional[str] = None  # "credito" | "debito" | "compensacao" | "ajuste"


@rh_router.post("/banco-horas/ajustes")
async def criar_ajuste_banco_horas(data: BancoHorasAjusteCreate):
    """Cria um ajuste manual no banco de horas de um funcionário.
    `minutos` positivo adiciona crédito, negativo registra débito.
    Use isso para acertos legais (compensação, decisão judicial, fechamento contábil)."""
    func = await funcionarios_collection.find_one({"id": data.funcionario_id}, {"_id": 0})
    if not func:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    if not data.motivo or not data.motivo.strip():
        raise HTTPException(status_code=400, detail="Motivo é obrigatório")
    if data.minutos == 0:
        raise HTTPException(status_code=400, detail="Quantidade de minutos não pode ser 0")

    doc = {
        "id": str(uuid.uuid4()),
        "funcionario_id": data.funcionario_id,
        "funcionario_nome": func.get("nome"),
        "minutos": int(data.minutos),
        "data": data.data or datetime.now().strftime("%Y-%m-%d"),
        "motivo": data.motivo.strip(),
        "tipo": data.tipo or ("credito" if data.minutos > 0 else "debito"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.banco_horas_ajustes.insert_one(doc)
    doc.pop("_id", None)
    return doc


@rh_router.get("/banco-horas/ajustes")
async def listar_ajustes_banco_horas(funcionario_id: Optional[str] = None):
    """Lista ajustes manuais (filtrado opcionalmente por funcionário)."""
    q: dict = {}
    if funcionario_id:
        q["funcionario_id"] = funcionario_id
    docs = []
    async for d in db.banco_horas_ajustes.find(q, {"_id": 0}).sort("data", -1):
        docs.append(d)
    return docs


@rh_router.delete("/banco-horas/ajustes/{ajuste_id}")
async def deletar_ajuste_banco_horas(ajuste_id: str):
    """Remove um ajuste manual de banco de horas."""
    res = await db.banco_horas_ajustes.delete_one({"id": ajuste_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ajuste não encontrado")
    return {"ok": True}


@rh_router.get("/banco-horas/funcionarios/{funcionario_id}/extrato")
async def banco_horas_extrato_funcionario(
    funcionario_id: str,
    ate_data: Optional[str] = None,
    de_data: Optional[str] = None,
):
    """Extrato detalhado do banco de horas de um funcionário no período."""
    func = await funcionarios_collection.find_one({"id": funcionario_id}, {"_id": 0})
    if not func:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")

    if ate_data is None:
        ate_data = datetime.now().strftime("%Y-%m-%d")

    # Resolve jornada
    jornada_padrao = await _get_or_create_jornada_padrao()
    jornada_func = jornada_padrao
    if func.get("jornada_id"):
        j_doc = await jornadas_collection.find_one({"id": func["jornada_id"]}, {"_id": 0})
        if j_doc:
            jornada_func = j_doc

    # Carrega abonos do funcionário
    abonos_data: dict = {}
    async for ab in ponto_abonos_collection.find({"funcionario_id": funcionario_id}, {"_id": 0}):
        abonos_data[ab["data"]] = ab

    # Carrega TODOS os pontos do funcionário no período
    data_filter: dict = {"$lte": ate_data}
    if de_data:
        data_filter["$gte"] = de_data
    pontos = []
    async for r in ponto_collection.find(
        {"funcionario_id": funcionario_id, "data": data_filter}, {"_id": 0}
    ):
        pontos.append(r)
    pontos.sort(key=lambda x: x["data"])

    # Carrega ajustes manuais do funcionário no período
    ajustes_filter: dict = {"funcionario_id": funcionario_id, "data": data_filter}
    ajustes_list = []
    async for aj in db.banco_horas_ajustes.find(ajustes_filter, {"_id": 0}).sort("data", 1):
        ajustes_list.append(aj)

    # Detalhe diário
    detalhe_dias = []
    saldo_acumulado = 0
    por_mes: dict = {}
    for r in pontos:
        prev = _jornada_minutos_previstos(jornada_func, r.get("dia_semana", 6))
        trab = int(r.get("minutos_trabalhados", 0) or 0)
        ab = abonos_data.get(r["data"])
        saldo_dia = 0 if ab else (trab - prev)
        saldo_acumulado += saldo_dia
        detalhe_dias.append({
            "data": r["data"],
            "dia_semana": r.get("dia_semana"),
            "batidas": r.get("batidas", []),
            "minutos_trabalhados": trab,
            "minutos_previstos": prev,
            "saldo_dia_minutos": saldo_dia,
            "saldo_acumulado_minutos": saldo_acumulado,
            "status_dia": "abonado" if ab else r.get("status_dia"),
            "abono": ({"tipo": ab["tipo"], "motivo": ab["motivo"]} if ab else None),
        })
        mes_chave = r["data"][:7]
        if mes_chave not in por_mes:
            por_mes[mes_chave] = {
                "mes": mes_chave,
                "minutos_trabalhados": 0,
                "minutos_previstos": 0,
                "saldo_minutos": 0,
                "dias": 0,
                "abonos": 0,
            }
        por_mes[mes_chave]["minutos_trabalhados"] += trab
        por_mes[mes_chave]["minutos_previstos"] += prev
        por_mes[mes_chave]["saldo_minutos"] += saldo_dia
        por_mes[mes_chave]["dias"] += 1
        if ab:
            por_mes[mes_chave]["abonos"] += 1

    # Aplica ajustes manuais ao saldo total e por mês
    total_ajustes = 0
    for aj in ajustes_list:
        delta = int(aj.get("minutos") or 0)
        saldo_acumulado += delta
        total_ajustes += delta
        mes_chave = (aj.get("data") or "")[:7]
        if mes_chave not in por_mes:
            por_mes[mes_chave] = {
                "mes": mes_chave, "minutos_trabalhados": 0, "minutos_previstos": 0,
                "saldo_minutos": 0, "dias": 0, "abonos": 0,
            }
        por_mes[mes_chave]["saldo_minutos"] += delta

    # Calcular saldo acumulado mês a mês
    evolucao = []
    saldo_corrente = 0
    for mes_chave in sorted(por_mes.keys()):
        item = por_mes[mes_chave]
        saldo_corrente += item["saldo_minutos"]
        evolucao.append({
            **item,
            "saldo_acumulado_minutos": saldo_corrente,
        })

    return {
        "funcionario": {
            "id": func["id"],
            "nome": func.get("nome", "-"),
            "cargo": func.get("cargo", "-"),
            "departamento": func.get("departamento", "-"),
            "data_admissao": func.get("data_admissao"),
            "jornada_nome": jornada_func.get("nome", "Padrão"),
        },
        "de_data": de_data,
        "ate_data": ate_data,
        "saldo_total_minutos": saldo_acumulado,
        "ajustes_minutos": total_ajustes,
        "total_dias": len(detalhe_dias),
        "total_abonos": sum(1 for d in detalhe_dias if d.get("abono")),
        "evolucao_mensal": evolucao,
        "detalhe_dias": detalhe_dias,
        "ajustes": ajustes_list,
    }


@rh_router.get("/banco-horas/funcionarios/{funcionario_id}/extrato-pdf")
async def exportar_extrato_banco_horas_pdf(
    funcionario_id: str,
    ate_data: Optional[str] = None,
    de_data: Optional[str] = None,
):
    """Exporta o extrato detalhado do banco de horas em PDF."""
    extrato = await banco_horas_extrato_funcionario(funcionario_id, ate_data, de_data)
    pdf_bytes = _build_extrato_banco_horas_pdf(extrato)
    func = extrato["funcionario"]
    nome_safe = (func.get("nome") or "func").replace(" ", "_")[:40]
    filename = f"BancoHoras_{nome_safe}_{extrato['ate_data']}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _build_extrato_banco_horas_pdf(extrato: dict) -> bytes:
    """Gera PDF do extrato de banco de horas com layout corporativo padronizado."""
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from utils.pdf_template import (
        create_corporate_doc, add_corporate_header, add_footer,
        get_corporate_styles, build_data_table, BRAND_COLORS, header_table_style,
    )

    func = extrato["funcionario"]
    saldo_total = extrato["saldo_total_minutos"]
    
    def _br_date(s: str) -> str:
        if not s:
            return "-"
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            return s

    def _mes_label(s: str) -> str:
        meses = ["", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                 "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
        try:
            ano, mes = s.split("-")
            return f"{meses[int(mes)]}/{ano}"
        except Exception:
            return s

    buffer = io.BytesIO()
    doc = create_corporate_doc(
        buffer,
        title=f"Banco de Horas - {func.get('nome','funcionario')}",
    )
    styles = get_corporate_styles()

    elements = []
    add_corporate_header(
        elements,
        doc_title="EXTRATO DE BANCO DE HORAS",
        subtitle=f"Apuração até {_br_date(extrato['ate_data'])}",
    )

    # Identificação
    elements.append(Paragraph("FUNCIONÁRIO", styles["section"]))
    elements.append(build_data_table([
        ("Nome:", func.get("nome", "-")),
        ("Cargo:", func.get("cargo", "-")),
        ("Departamento:", func.get("departamento", "-")),
        ("Admissão:", _br_date(func.get("data_admissao", ""))),
        ("Jornada:", func.get("jornada_nome", "Padrão")),
    ]))
    elements.append(Spacer(1, 14))

    # Saldo destacado
    cor_destaque = BRAND_COLORS["primary"] if saldo_total >= 0 else colors.HexColor("#dc2626")
    label_saldo = "SALDO POSITIVO (CRÉDITO)" if saldo_total >= 0 else "SALDO NEGATIVO (DÉBITO)"
    t_saldo = Table([
        [label_saldo, _fmt_min_pdf(saldo_total)],
    ], colWidths=[10 * cm, 6 * cm])
    t_saldo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), cor_destaque),
        ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
        ("BACKGROUND", (1, 0), (1, 0), BRAND_COLORS["accent"]),
        ("TEXTCOLOR", (1, 0), (1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 14),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
    ]))
    elements.append(t_saldo)
    elements.append(Spacer(1, 16))

    # Evolução mensal
    elements.append(Paragraph("EVOLUÇÃO MENSAL", styles["section"]))
    rows = [["Mês", "Trabalhado", "Previsto", "Saldo Mês", "Saldo Acumulado", "Dias", "Abonos"]]
    for ev in extrato.get("evolucao_mensal", []):
        rows.append([
            _mes_label(ev["mes"]),
            _fmt_min_pdf(ev["minutos_trabalhados"]),
            _fmt_min_pdf(ev["minutos_previstos"]),
            _fmt_min_pdf(ev["saldo_minutos"]),
            _fmt_min_pdf(ev["saldo_acumulado_minutos"]),
            str(ev["dias"]),
            str(ev["abonos"]),
        ])
    t_evo = Table(
        rows,
        colWidths=[2.4 * cm, 2.7 * cm, 2.7 * cm, 2.7 * cm, 3.0 * cm, 1.6 * cm, 1.8 * cm],
        repeatRows=1,
    )
    style = header_table_style()
    style.add("ALIGN", (1, 1), (-1, -1), "CENTER")
    t_evo.setStyle(style)
    elements.append(t_evo)
    elements.append(Spacer(1, 14))

    # Detalhe diário (limitar a últimos 90 dias para não estourar)
    detalhe = extrato.get("detalhe_dias", [])[-90:]
    if detalhe:
        elements.append(Paragraph(
            f"DETALHE DIÁRIO (últimos {len(detalhe)} registros)",
            styles["section"],
        ))
        rows = [["Data", "Dia", "Batidas", "Trab.", "Previsto", "Saldo Dia", "Acum.", "Status"]]
        dias_pt = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
        for d in detalhe:
            ds = d.get("dia_semana")
            dia_lbl = dias_pt[ds] if isinstance(ds, int) and 0 <= ds < 7 else "-"
            batidas_str = " ".join(d.get("batidas") or []) or "-"
            status = d.get("status_dia") or "-"
            rows.append([
                _br_date(d.get("data", "")),
                dia_lbl,
                batidas_str[:30],
                _fmt_min_pdf(d.get("minutos_trabalhados", 0)),
                _fmt_min_pdf(d.get("minutos_previstos", 0)),
                _fmt_min_pdf(d.get("saldo_dia_minutos", 0)),
                _fmt_min_pdf(d.get("saldo_acumulado_minutos", 0)),
                status,
            ])
        t_det = Table(
            rows,
            colWidths=[2.0 * cm, 1.2 * cm, 3.5 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 2.0 * cm, 2.0 * cm],
            repeatRows=1,
        )
        style = header_table_style()
        style.add("ALIGN", (1, 1), (-1, -1), "CENTER")
        style.add("FONTSIZE", (0, 0), (-1, -1), 7)
        t_det.setStyle(style)
        elements.append(t_det)

    add_footer(elements, "Sistema CRA · Extrato de Banco de Horas")
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


# ===== PDF: ESPELHO DE PONTO MENSAL =====
def _fmt_min_pdf(min_total: int) -> str:
    """Formata minutos como '+12h 30min' / '-3h 45min' / '0h'"""
    if min_total == 0:
        return "0h"
    sinal = "-" if min_total < 0 else ""
    abs_min = abs(min_total)
    h = abs_min // 60
    m = abs_min % 60
    if m == 0:
        return f"{sinal}{h}h"
    return f"{sinal}{h}h {m}min"


def _build_espelho_ponto_pdf(funcionarios: list, mes: int, ano: int) -> bytes:
    """Gera PDF de Espelho de Ponto. Aceita 1 ou N funcionários (cada um em uma seção)."""
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, Spacer, Table, TableStyle, PageBreak,
    )
    from utils.pdf_template import (
        create_corporate_doc, add_corporate_header, add_footer, BRAND_COLORS,
    )
    
    meses_pt = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    dias_semana_pt = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    
    buffer = io.BytesIO()
    doc = create_corporate_doc(buffer, title=f"Espelho de Ponto {meses_pt[mes]}/{ano}")
    styles = getSampleStyleSheet()
    section_style = ParagraphStyle("Section", parent=styles["Heading4"],
                                   fontSize=10, spaceBefore=8, spaceAfter=4,
                                   textColor=BRAND_COLORS["primary"])
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=7,
                           textColor=BRAND_COLORS["muted"])
    
    elements = []
    for idx, f in enumerate(funcionarios):
        if idx > 0:
            elements.append(PageBreak())
        
        # CABEÇALHO CORPORATIVO (logo + título + subtítulo + linha divisória)
        add_corporate_header(
            elements,
            doc_title="ESPELHO DE PONTO ELETRÔNICO",
            subtitle=f"Competência: {meses_pt[mes]} / {ano}",
        )
        
        # Bloco de identificação do funcionário
        identif_data = [
            ["Funcionário:", f.get("nome", "-"), "Cargo:", f.get("cargo", "-")],
            ["Departamento:", f.get("departamento", "-"), "Status:",
             "Cadastrado" if f.get("cadastrado") else "NÃO CADASTRADO NA PLATAFORMA"],
        ]
        t_identif = Table(identif_data, colWidths=[2.5 * cm, 7 * cm, 2 * cm, 6.5 * cm])
        t_identif.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ]))
        elements.append(t_identif)
        elements.append(Spacer(1, 6))
        
        # Resumo do mês (4 quadros lado a lado)
        elements.append(Paragraph("RESUMO DO PERÍODO", section_style))
        
        saldo_minutos = f.get("saldo_mes_minutos", 0)
        banco_minutos = f.get("banco_horas_acumulado_minutos", 0)
        saldo_str = ("+" if saldo_minutos > 0 else "") + _fmt_min_pdf(saldo_minutos)
        banco_str = ("+" if banco_minutos > 0 else "") + _fmt_min_pdf(banco_minutos)
        
        resumo_data = [
            ["Trabalhadas", "Previstas", "Saldo do mês", "Banco acumulado"],
            [
                _fmt_min_pdf(f.get("minutos_trabalhados", 0)),
                _fmt_min_pdf(f.get("minutos_previstos", 0)),
                saldo_str,
                banco_str,
            ],
        ]
        t_resumo = Table(resumo_data, colWidths=[4.5 * cm, 4.5 * cm, 4.5 * cm, 4.5 * cm])
        cor_saldo = colors.HexColor("#10B981") if saldo_minutos >= 0 else colors.HexColor("#ef4444")
        cor_banco = colors.HexColor("#2563eb") if banco_minutos >= 0 else colors.HexColor("#ea580c")
        t_resumo.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, 1), 12),
            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("TEXTCOLOR", (0, 1), (0, 1), colors.HexColor("#10B981")),
            ("TEXTCOLOR", (1, 1), (1, 1), colors.HexColor("#475569")),
            ("TEXTCOLOR", (2, 1), (2, 1), cor_saldo),
            ("TEXTCOLOR", (3, 1), (3, 1), cor_banco),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(t_resumo)
        
        # Estatísticas de dias
        stats_data = [
            ["Dias trabalhados", "Dias incompletos", "Faltas"],
            [
                str(f.get("dias_com_registro", 0)),
                str(f.get("dias_incompletos", 0)),
                str(f.get("dias_falta", 0)),
            ],
        ]
        t_stats = Table(stats_data, colWidths=[6 * cm, 6 * cm, 6 * cm])
        t_stats.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7),
            ("FONTSIZE", (0, 1), (-1, 1), 11),
            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#dcfce7")),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#fef3c7")),
            ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#fee2e2")),
            ("TEXTCOLOR", (0, 1), (0, 1), colors.HexColor("#15803d")),
            ("TEXTCOLOR", (1, 1), (1, 1), colors.HexColor("#a16207")),
            ("TEXTCOLOR", (2, 1), (2, 1), colors.HexColor("#b91c1c")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(Spacer(1, 4))
        elements.append(t_stats)
        elements.append(Spacer(1, 8))
        
        # Tabela detalhada dia a dia
        elements.append(Paragraph("DETALHAMENTO DIÁRIO", section_style))
        
        cab = ["Data", "Dia", "Batidas registradas", "Trabalhado", "Previsto", "Saldo"]
        linhas = [cab]
        
        for d in f.get("detalhe_dias", []):
            data = d.get("data", "")
            data_br = "/".join(reversed(data.split("-"))) if data else "-"
            dia_idx = d.get("dia_semana", 0)
            dia_lbl = dias_semana_pt[dia_idx] if 0 <= dia_idx < 7 else "?"
            batidas_lst = d.get("batidas") or []
            batidas_str = " | ".join(batidas_lst) if batidas_lst else "—"
            saldo = d.get("saldo_minutos", 0)
            saldo_lbl = ("+" if saldo > 0 else "") + _fmt_min_pdf(saldo)
            
            # Marcador de abono na coluna de batidas
            ab = d.get("abono")
            if ab:
                tipo_lbl = (ab.get("tipo") or "abono").upper()
                motivo_lbl = ab.get("motivo") or ""
                batidas_str = f"[ABONO {tipo_lbl}] {motivo_lbl}"[:80]
                saldo_lbl = "ABONADO"
            
            linhas.append([
                data_br,
                dia_lbl,
                batidas_str,
                _fmt_min_pdf(d.get("minutos_trabalhados", 0)),
                _fmt_min_pdf(d.get("minutos_previstos", 0)),
                saldo_lbl,
            ])
        
        t_dias = Table(linhas, colWidths=[2.2 * cm, 1.5 * cm, 6.8 * cm, 2.4 * cm, 2.4 * cm, 2.7 * cm])
        ts = TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (2, 1), (2, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ])
        # Cores por status: saldo positivo verde, negativo vermelho, abonado amarelo
        for i, d in enumerate(f.get("detalhe_dias", []), start=1):
            ab = d.get("abono")
            if ab:
                ts.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#fef3c7"))
                ts.add("TEXTCOLOR", (5, i), (5, i), colors.HexColor("#a16207"))
                ts.add("FONTNAME", (5, i), (5, i), "Helvetica-Bold")
            else:
                saldo = d.get("saldo_minutos", 0)
                if saldo > 0:
                    ts.add("TEXTCOLOR", (5, i), (5, i), colors.HexColor("#10B981"))
                elif saldo < 0:
                    ts.add("TEXTCOLOR", (5, i), (5, i), colors.HexColor("#dc2626"))
        t_dias.setStyle(ts)
        elements.append(t_dias)
        elements.append(Spacer(1, 8))
        
        # Lista detalhada de abonos
        abonos = f.get("abonos") or []
        if abonos:
            elements.append(Paragraph("ABONOS DO MÊS", section_style))
            ab_cab = ["Data", "Tipo", "Motivo / Justificativa", "Anexo"]
            ab_linhas = [ab_cab]
            for ab in sorted(abonos, key=lambda x: x.get("data", "")):
                d_br = "/".join(reversed(ab.get("data", "").split("-")))
                anexo = ab.get("anexo") or {}
                anexo_str = "Sim" if anexo.get("storage_path") else "—"
                if anexo.get("filename_original"):
                    anexo_str = f"Sim ({anexo['filename_original'][:24]})"
                ab_linhas.append([
                    d_br,
                    (ab.get("tipo") or "").upper(),
                    ab.get("motivo") or "",
                    anexo_str,
                ])
            t_ab = Table(ab_linhas, colWidths=[2.2 * cm, 2.6 * cm, 9.5 * cm, 3.7 * cm])
            t_ab.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#a16207")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (2, 1), (2, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(t_ab)
            elements.append(Spacer(1, 8))
        
        # Observações livres do mês
        obs_texto = (f.get("observacao") or "").strip()
        if obs_texto:
            elements.append(Paragraph("OBSERVAÇÕES", section_style))
            obs_para = Paragraph(obs_texto.replace("\n", "<br/>"), 
                                 ParagraphStyle("Obs", parent=styles["Normal"], fontSize=8.5, leading=11))
            t_obs = Table([[obs_para]], colWidths=[18 * cm])
            t_obs.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fffbeb")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#fbbf24")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(t_obs)
            elements.append(Spacer(1, 12))
        else:
            elements.append(Spacer(1, 6))
        
        # Linhas de assinatura
        elements.append(Paragraph(
            "Declaro que as informações registradas neste espelho de ponto estão corretas e foram conferidas.",
            small,
        ))
        elements.append(Spacer(1, 24))
        
        sign_data = [
            ["_" * 40, "", "_" * 40],
            ["Assinatura do Funcionário", "", "Assinatura do Responsável"],
        ]
        t_sign = Table(sign_data, colWidths=[7 * cm, 1 * cm, 7 * cm])
        t_sign.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#475569")),
        ]))
        elements.append(t_sign)
        elements.append(Spacer(1, 8))
    
    # Rodapé corporativo
    add_footer(elements, "Sistema CRA · Espelho de Ponto Eletrônico")
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


@rh_router.get("/ponto/relatorio-pdf")
async def gerar_pdf_espelho_ponto(
    mes: int,
    ano: int,
    funcionario_id: Optional[str] = None,
):
    """Gera espelho de ponto em PDF.
    Se `funcionario_id` for informado, retorna PDF de 1 funcionário.
    Caso contrário, retorna PDF consolidado com todos os funcionários do mês."""
    # Reusa a lógica do dashboard_mensal para garantir cálculos idênticos
    dashboard = await get_ponto_dashboard_mensal(mes=mes, ano=ano)
    
    funcionarios = dashboard.get("funcionarios", [])
    if funcionario_id:
        funcionarios = [f for f in funcionarios if f.get("funcionario_id") == funcionario_id]
        if not funcionarios:
            raise HTTPException(status_code=404, detail="Funcionário não encontrado neste mês")
    
    if not funcionarios:
        raise HTTPException(status_code=404, detail="Nenhum registro de ponto encontrado para este período")
    
    pdf_bytes = _build_espelho_ponto_pdf(funcionarios, mes, ano)
    
    if funcionario_id:
        nome_seguro = (funcionarios[0].get("nome") or "funcionario").replace(" ", "_")[:40]
        filename = f"EspelhoPonto_{nome_seguro}_{mes:02d}_{ano}.pdf"
    else:
        filename = f"EspelhoPonto_TODOS_{mes:02d}_{ano}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )



# ===== JORNADAS DE TRABALHO PERSONALIZADAS =====
DIAS_SEMANA_PT = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
JORNADA_PADRAO_NOME = "Padrão"


def _hhmm_to_min(s: str) -> int:
    """Converte 'HH:MM' em minutos. Retorna 0 se inválido/vazio."""
    if not s:
        return 0
    try:
        h, m = s.split(":")
        return int(h) * 60 + int(m)
    except (ValueError, AttributeError):
        return 0


def _calc_minutos_dia_jornada(dia_cfg: dict) -> int:
    """Calcula minutos previstos para um dia da jornada.
    dia_cfg = { ativo, entrada, saida_almoco, retorno_almoco, saida }
    Se almoço presente: (saida_almoco - entrada) + (saida - retorno_almoco)
    Se sem almoço: saida - entrada
    Retorna 0 se inativo ou inválido.
    """
    if not dia_cfg or not dia_cfg.get("ativo"):
        return 0
    e = _hhmm_to_min(dia_cfg.get("entrada"))
    s = _hhmm_to_min(dia_cfg.get("saida"))
    sa = _hhmm_to_min(dia_cfg.get("saida_almoco"))
    ra = _hhmm_to_min(dia_cfg.get("retorno_almoco"))
    if e <= 0 or s <= 0 or s <= e:
        return 0
    if sa > 0 and ra > 0 and sa < ra:
        # Com almoço
        return max(0, (sa - e) + (s - ra))
    # Sem almoço
    return max(0, s - e)


def _jornada_minutos_previstos(jornada_doc: Optional[dict], dia_semana: int) -> int:
    """Retorna minutos previstos para um dia_semana (0=Seg, 6=Dom) baseado na jornada.
    Se jornada_doc é None, usa fallback hardcoded (Seg-Sex 8h, Sáb 4h, Dom 0h)."""
    if jornada_doc and isinstance(jornada_doc.get("dias"), dict):
        dia_cfg = jornada_doc["dias"].get(str(dia_semana))
        return _calc_minutos_dia_jornada(dia_cfg or {})
    # Fallback
    if 0 <= dia_semana <= 4:
        return 8 * 60
    if dia_semana == 5:
        return 4 * 60
    return 0


def _jornada_total_semanal_min(jornada_doc: dict) -> int:
    if not jornada_doc or not isinstance(jornada_doc.get("dias"), dict):
        return 0
    total = 0
    for d in range(7):
        total += _calc_minutos_dia_jornada(jornada_doc["dias"].get(str(d)) or {})
    return total


async def _get_or_create_jornada_padrao() -> dict:
    """Retorna a jornada Padrão, criando-a se não existir."""
    j = await jornadas_collection.find_one({"nome": JORNADA_PADRAO_NOME}, {"_id": 0})
    if j:
        return j
    novo = {
        "id": str(uuid.uuid4()),
        "nome": JORNADA_PADRAO_NOME,
        "descricao": "Jornada padrão: Seg-Sex 08:00-17:00 (1h almoço) e Sábado 08:00-12:00",
        "is_padrao": True,
        "dias": {
            "0": {"ativo": True, "entrada": "08:00", "saida_almoco": "11:30",
                  "retorno_almoco": "13:30", "saida": "18:00"},
            "1": {"ativo": True, "entrada": "08:00", "saida_almoco": "11:30",
                  "retorno_almoco": "13:30", "saida": "18:00"},
            "2": {"ativo": True, "entrada": "08:00", "saida_almoco": "11:30",
                  "retorno_almoco": "13:30", "saida": "18:00"},
            "3": {"ativo": True, "entrada": "08:00", "saida_almoco": "11:30",
                  "retorno_almoco": "13:30", "saida": "18:00"},
            "4": {"ativo": True, "entrada": "08:00", "saida_almoco": "11:30",
                  "retorno_almoco": "13:30", "saida": "18:00"},
            "5": {"ativo": True, "entrada": "08:00", "saida_almoco": "",
                  "retorno_almoco": "", "saida": "12:00"},
            "6": {"ativo": False, "entrada": "", "saida_almoco": "",
                  "retorno_almoco": "", "saida": ""},
        },
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await jornadas_collection.insert_one(novo)
    novo.pop("_id", None)
    return novo


@rh_router.get("/jornadas")
async def listar_jornadas():
    """Lista todas as jornadas com contagem de funcionários atribuídos."""
    await _get_or_create_jornada_padrao()
    jornadas = []
    async for j in jornadas_collection.find({}, {"_id": 0}).sort("nome", 1):
        # Conta funcionários atribuídos
        count = await funcionarios_collection.count_documents({"jornada_id": j["id"]})
        # Inclui contagem de funcionários SEM jornada que receberão a padrão
        if j.get("is_padrao"):
            sem_jornada = await funcionarios_collection.count_documents(
                {"$or": [{"jornada_id": {"$exists": False}}, {"jornada_id": None}, {"jornada_id": ""}]}
            )
            count += sem_jornada
        j["funcionarios_count"] = count
        j["total_semanal_minutos"] = _jornada_total_semanal_min(j)
        jornadas.append(j)
    return jornadas


@rh_router.post("/jornadas")
async def criar_jornada(payload: dict = Body(...)):
    nome = (payload.get("nome") or "").strip()
    descricao = (payload.get("descricao") or "").strip()
    dias = payload.get("dias") or {}
    
    if not nome:
        raise HTTPException(status_code=400, detail="Nome é obrigatório")
    if await jornadas_collection.find_one({"nome": nome}):
        raise HTTPException(status_code=400, detail="Já existe jornada com este nome")
    
    # Normaliza dias (chaves 0..6 como string)
    dias_norm = {}
    for d in range(7):
        cfg = dias.get(str(d)) or dias.get(d) or {}
        dias_norm[str(d)] = {
            "ativo": bool(cfg.get("ativo", False)),
            "entrada": (cfg.get("entrada") or "").strip(),
            "saida_almoco": (cfg.get("saida_almoco") or "").strip(),
            "retorno_almoco": (cfg.get("retorno_almoco") or "").strip(),
            "saida": (cfg.get("saida") or "").strip(),
        }
    
    doc = {
        "id": str(uuid.uuid4()),
        "nome": nome,
        "descricao": descricao,
        "is_padrao": False,
        "dias": dias_norm,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await jornadas_collection.insert_one(doc)
    doc.pop("_id", None)
    doc["funcionarios_count"] = 0
    doc["total_semanal_minutos"] = _jornada_total_semanal_min(doc)
    return doc


@rh_router.put("/jornadas/{jornada_id}")
async def atualizar_jornada(jornada_id: str, payload: dict = Body(...)):
    existing = await jornadas_collection.find_one({"id": jornada_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Jornada não encontrada")
    
    update = {}
    if "nome" in payload:
        novo_nome = (payload.get("nome") or "").strip()
        if not novo_nome:
            raise HTTPException(status_code=400, detail="Nome não pode ser vazio")
        outro = await jornadas_collection.find_one({"nome": novo_nome, "id": {"$ne": jornada_id}})
        if outro:
            raise HTTPException(status_code=400, detail="Já existe outra jornada com este nome")
        update["nome"] = novo_nome
    if "descricao" in payload:
        update["descricao"] = (payload.get("descricao") or "").strip()
    if "dias" in payload:
        dias_norm = {}
        for d in range(7):
            cfg = payload["dias"].get(str(d)) or payload["dias"].get(d) or {}
            dias_norm[str(d)] = {
                "ativo": bool(cfg.get("ativo", False)),
                "entrada": (cfg.get("entrada") or "").strip(),
                "saida_almoco": (cfg.get("saida_almoco") or "").strip(),
                "retorno_almoco": (cfg.get("retorno_almoco") or "").strip(),
                "saida": (cfg.get("saida") or "").strip(),
            }
        update["dias"] = dias_norm
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await jornadas_collection.update_one({"id": jornada_id}, {"$set": update})
    j = await jornadas_collection.find_one({"id": jornada_id}, {"_id": 0})
    j["funcionarios_count"] = await funcionarios_collection.count_documents({"jornada_id": jornada_id})
    j["total_semanal_minutos"] = _jornada_total_semanal_min(j)
    return j


@rh_router.delete("/jornadas/{jornada_id}")
async def deletar_jornada(jornada_id: str):
    j = await jornadas_collection.find_one({"id": jornada_id})
    if not j:
        raise HTTPException(status_code=404, detail="Jornada não encontrada")
    if j.get("is_padrao"):
        raise HTTPException(status_code=400, detail="Não é possível excluir a jornada Padrão")
    count = await funcionarios_collection.count_documents({"jornada_id": jornada_id})
    if count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Não é possível excluir: {count} funcionário(s) ainda usam esta jornada. Remova ou troque a jornada deles primeiro.",
        )
    await jornadas_collection.delete_one({"id": jornada_id})
    return {"message": "Jornada removida"}


@rh_router.post("/jornadas/{jornada_id}/atribuir")
async def atribuir_funcionarios_jornada(jornada_id: str, payload: dict = Body(...)):
    """Atribui múltiplos funcionários a uma jornada.
    Body: { funcionario_ids: [...] }
    Substitui as jornadas atuais desses funcionários."""
    j = await jornadas_collection.find_one({"id": jornada_id})
    if not j:
        raise HTTPException(status_code=404, detail="Jornada não encontrada")
    func_ids = payload.get("funcionario_ids") or []
    if not isinstance(func_ids, list):
        raise HTTPException(status_code=400, detail="funcionario_ids deve ser uma lista")
    
    res = await funcionarios_collection.update_many(
        {"id": {"$in": func_ids}},
        {"$set": {"jornada_id": jornada_id}},
    )
    return {"message": f"{res.modified_count} funcionário(s) atribuído(s) à jornada", "modified": res.modified_count}


@rh_router.get("/jornadas/{jornada_id}/funcionarios")
async def listar_funcionarios_da_jornada(jornada_id: str):
    """Lista funcionários atribuídos à jornada. Para a 'Padrão', inclui também os sem jornada definida."""
    j = await jornadas_collection.find_one({"id": jornada_id})
    if not j:
        raise HTTPException(status_code=404, detail="Jornada não encontrada")
    
    if j.get("is_padrao"):
        query = {"$or": [
            {"jornada_id": jornada_id},
            {"jornada_id": {"$exists": False}},
            {"jornada_id": None},
            {"jornada_id": ""},
        ]}
    else:
        query = {"jornada_id": jornada_id}
    
    funcs = []
    async for f in funcionarios_collection.find(query, {"_id": 0, "id": 1, "nome": 1, "cargo": 1, "departamento": 1}).sort("nome", 1):
        funcs.append(f)
    return funcs


# ===== ABONOS DE PONTO (faltas justificadas / atestados / folgas / feriados) =====
TIPOS_ABONO_VALIDOS = {"atestado", "justificativa", "feriado", "folga", "ferias", "outros"}
ABONO_EXTS_VALIDAS = {"pdf", "jpg", "jpeg", "png", "webp", "heic", "heif"}
ABONO_MAX_BYTES = 10 * 1024 * 1024  # 10 MB


def _ext_arquivo(filename: str) -> str:
    if not filename or "." not in filename:
        return ""
    return filename.rsplit(".", 1)[-1].lower().strip()


@rh_router.post("/ponto/abono")
async def criar_abono(
    funcionario_id: str = Body(...),
    data: str = Body(...),
    tipo: str = Body(...),
    motivo: str = Body(...),
):
    """Cria um abono SEM anexo. Body JSON: { funcionario_id, data, tipo, motivo }."""
    funcionario_id = (funcionario_id or "").strip()
    data = (data or "").strip()
    tipo = (tipo or "").strip().lower()
    motivo = (motivo or "").strip()
    
    if not funcionario_id:
        raise HTTPException(status_code=400, detail="funcionario_id é obrigatório")
    if not data:
        raise HTTPException(status_code=400, detail="data é obrigatória (YYYY-MM-DD)")
    try:
        datetime.strptime(data, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="data inválida (use YYYY-MM-DD)")
    if tipo not in TIPOS_ABONO_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail=f"tipo inválido. Use um de: {', '.join(sorted(TIPOS_ABONO_VALIDOS))}",
        )
    if not motivo:
        raise HTTPException(status_code=400, detail="motivo é obrigatório")
    
    await ponto_abonos_collection.delete_many({
        "funcionario_id": funcionario_id,
        "data": data,
    })
    
    doc = {
        "id": str(uuid.uuid4()),
        "funcionario_id": funcionario_id,
        "data": data,
        "tipo": tipo,
        "motivo": motivo,
        "anexo": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await ponto_abonos_collection.insert_one(doc)
    doc.pop("_id", None)
    return doc


@rh_router.post("/ponto/abono-com-anexo")
async def criar_abono_com_anexo(
    funcionario_id: str = Form(...),
    data: str = Form(...),
    tipo: str = Form(...),
    motivo: str = Form(...),
    arquivo: Optional[UploadFile] = File(None),
):
    """Cria abono e (opcional) faz upload de atestado/justificativa.
    Multipart/form-data. Aceita PDF/JPG/PNG/WEBP/HEIC até 10MB."""
    from fastapi import Form  # noqa: F401 (já importado no topo via UploadFile/File)
    funcionario_id = (funcionario_id or "").strip()
    data_str = (data or "").strip()
    tipo = (tipo or "").strip().lower()
    motivo = (motivo or "").strip()
    
    if not funcionario_id:
        raise HTTPException(status_code=400, detail="funcionario_id é obrigatório")
    if not data_str:
        raise HTTPException(status_code=400, detail="data é obrigatória")
    try:
        datetime.strptime(data_str, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="data inválida (YYYY-MM-DD)")
    if tipo not in TIPOS_ABONO_VALIDOS:
        raise HTTPException(status_code=400, detail="tipo inválido")
    if not motivo:
        raise HTTPException(status_code=400, detail="motivo é obrigatório")
    
    anexo_doc = None
    if arquivo and arquivo.filename:
        ext = _ext_arquivo(arquivo.filename)
        if ext not in ABONO_EXTS_VALIDAS:
            raise HTTPException(
                status_code=400,
                detail=f"Extensão de arquivo não permitida ({ext}). Use: {', '.join(sorted(ABONO_EXTS_VALIDAS))}",
            )
        conteudo = await arquivo.read()
        if not conteudo:
            raise HTTPException(status_code=400, detail="Arquivo vazio")
        if len(conteudo) > ABONO_MAX_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"Arquivo muito grande ({len(conteudo) // (1024 * 1024)}MB). Máx: 10MB",
            )
        from utils.storage import put_object, MIME_BY_EXT, APP_NAME
        path = f"{APP_NAME}/abonos/{funcionario_id}/{uuid.uuid4()}.{ext}"
        content_type = MIME_BY_EXT.get(ext, arquivo.content_type or "application/octet-stream")
        try:
            result = put_object(path, conteudo, content_type)
            anexo_doc = {
                "storage_path": result["path"],
                "filename_original": arquivo.filename,
                "content_type": content_type,
                "size": result.get("size", len(conteudo)),
                "ext": ext,
            }
        except Exception as e:
            raise HTTPException(
                status_code=502,
                detail=f"Falha ao subir arquivo para o storage: {e}",
            )
    
    # Substitui abono anterior do mesmo (funcionario, data) — mas mantém anexo antigo se quiser preservar
    await ponto_abonos_collection.delete_many({
        "funcionario_id": funcionario_id,
        "data": data_str,
    })
    
    doc = {
        "id": str(uuid.uuid4()),
        "funcionario_id": funcionario_id,
        "data": data_str,
        "tipo": tipo,
        "motivo": motivo,
        "anexo": anexo_doc,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await ponto_abonos_collection.insert_one(doc)
    doc.pop("_id", None)
    return doc


@rh_router.post("/ponto/abono-em-massa")
async def criar_abono_em_massa(payload: dict = Body(...)):
    """Cria abonos em massa para um funcionário em várias datas.
    Body: { funcionario_id, datas: [str YYYY-MM-DD], tipo, motivo, anexo_storage_path?, anexo_meta? }
    Substitui qualquer abono existente nos mesmos (funcionario, data).
    Retorna: { criados: int, datas: [str], abonos: [...] }"""
    funcionario_id = (payload.get("funcionario_id") or "").strip()
    datas_raw = payload.get("datas") or []
    tipo = (payload.get("tipo") or "").strip().lower()
    motivo = (payload.get("motivo") or "").strip()
    anexo_doc = payload.get("anexo") or None  # opcional: já uploadado por outra rota

    if not funcionario_id:
        raise HTTPException(status_code=400, detail="funcionario_id é obrigatório")
    if not isinstance(datas_raw, list) or not datas_raw:
        raise HTTPException(status_code=400, detail="datas deve ser uma lista não vazia")
    if tipo not in TIPOS_ABONO_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail=f"tipo inválido. Use um de: {', '.join(sorted(TIPOS_ABONO_VALIDOS))}",
        )
    if not motivo:
        raise HTTPException(status_code=400, detail="motivo é obrigatório")

    # Valida e deduplica datas
    datas_validas = []
    seen = set()
    for d in datas_raw:
        d = (d or "").strip()
        if not d or d in seen:
            continue
        try:
            datetime.strptime(d, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail=f"data inválida: {d}")
        datas_validas.append(d)
        seen.add(d)

    if not datas_validas:
        raise HTTPException(status_code=400, detail="Nenhuma data válida")

    # Remove abonos existentes nas mesmas datas
    await ponto_abonos_collection.delete_many({
        "funcionario_id": funcionario_id,
        "data": {"$in": datas_validas},
    })

    docs = []
    now_iso = datetime.now(timezone.utc).isoformat()
    for d in datas_validas:
        docs.append({
            "id": str(uuid.uuid4()),
            "funcionario_id": funcionario_id,
            "data": d,
            "tipo": tipo,
            "motivo": motivo,
            "anexo": anexo_doc,
            "created_at": now_iso,
        })
    await ponto_abonos_collection.insert_many(docs)
    for doc in docs:
        doc.pop("_id", None)
    return {
        "criados": len(docs),
        "datas": datas_validas,
        "abonos": docs,
    }


@rh_router.post("/ponto/abono-em-massa-com-anexo")
async def criar_abono_em_massa_com_anexo(
    funcionario_id: str = Form(...),
    datas: str = Form(...),  # JSON-encoded list de strings YYYY-MM-DD
    tipo: str = Form(...),
    motivo: str = Form(...),
    arquivo: Optional[UploadFile] = File(None),
):
    """Versão multipart com anexo opcional COMPARTILHADO entre todas as datas.
    `datas` deve ser uma string JSON (ex: '["2026-05-01","2026-05-02"]')."""
    funcionario_id = (funcionario_id or "").strip()
    tipo_l = (tipo or "").strip().lower()
    motivo_s = (motivo or "").strip()
    if not funcionario_id:
        raise HTTPException(status_code=400, detail="funcionario_id é obrigatório")
    if tipo_l not in TIPOS_ABONO_VALIDOS:
        raise HTTPException(status_code=400, detail="tipo inválido")
    if not motivo_s:
        raise HTTPException(status_code=400, detail="motivo é obrigatório")
    try:
        datas_list = json.loads(datas) if isinstance(datas, str) else datas
    except Exception:
        raise HTTPException(status_code=400, detail="datas deve ser JSON: ['YYYY-MM-DD',...]")
    if not isinstance(datas_list, list) or not datas_list:
        raise HTTPException(status_code=400, detail="datas deve ser lista não vazia")

    datas_validas = []
    seen = set()
    for d in datas_list:
        d = (d or "").strip()
        if not d or d in seen:
            continue
        try:
            datetime.strptime(d, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail=f"data inválida: {d}")
        datas_validas.append(d)
        seen.add(d)
    if not datas_validas:
        raise HTTPException(status_code=400, detail="Nenhuma data válida")

    # Upload do anexo (uma vez, compartilhado entre todos)
    anexo_doc = None
    if arquivo and arquivo.filename:
        ext = _ext_arquivo(arquivo.filename)
        if ext not in ABONO_EXTS_VALIDAS:
            raise HTTPException(
                status_code=400,
                detail=f"Extensão não permitida ({ext}). Use: {', '.join(sorted(ABONO_EXTS_VALIDAS))}",
            )
        conteudo = await arquivo.read()
        if not conteudo:
            raise HTTPException(status_code=400, detail="Arquivo vazio")
        if len(conteudo) > ABONO_MAX_BYTES:
            raise HTTPException(status_code=400, detail="Arquivo > 10MB")
        from utils.storage import put_object, MIME_BY_EXT, APP_NAME
        path = f"{APP_NAME}/abonos/{funcionario_id}/{uuid.uuid4()}.{ext}"
        content_type = MIME_BY_EXT.get(ext, arquivo.content_type or "application/octet-stream")
        try:
            result = put_object(path, conteudo, content_type)
            anexo_doc = {
                "storage_path": result["path"],
                "filename_original": arquivo.filename,
                "content_type": content_type,
                "size": result.get("size", len(conteudo)),
                "ext": ext,
            }
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Falha no upload: {e}")

    await ponto_abonos_collection.delete_many({
        "funcionario_id": funcionario_id,
        "data": {"$in": datas_validas},
    })

    now_iso = datetime.now(timezone.utc).isoformat()
    docs = []
    for d in datas_validas:
        docs.append({
            "id": str(uuid.uuid4()),
            "funcionario_id": funcionario_id,
            "data": d,
            "tipo": tipo_l,
            "motivo": motivo_s,
            "anexo": anexo_doc,
            "created_at": now_iso,
        })
    await ponto_abonos_collection.insert_many(docs)
    for doc in docs:
        doc.pop("_id", None)
    return {
        "criados": len(docs),
        "datas": datas_validas,
        "anexo_compartilhado": bool(anexo_doc),
        "abonos": docs,
    }


@rh_router.get("/ponto/abono/{abono_id}/anexo")
async def baixar_anexo_abono(abono_id: str):
    """Baixa o arquivo anexo de um abono (PDF/imagem)."""
    abono = await ponto_abonos_collection.find_one({"id": abono_id}, {"_id": 0})
    if not abono:
        raise HTTPException(status_code=404, detail="Abono não encontrado")
    anexo = abono.get("anexo")
    if not anexo or not anexo.get("storage_path"):
        raise HTTPException(status_code=404, detail="Este abono não tem anexo")
    from utils.storage import get_object
    try:
        data, content_type = get_object(anexo["storage_path"])
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Erro ao baixar do storage: {e}")
    
    filename = anexo.get("filename_original") or f"anexo_abono.{anexo.get('ext', 'bin')}"
    return Response(
        content=data,
        media_type=anexo.get("content_type") or content_type,
        headers={"Content-Disposition": f"inline; filename=\"{filename}\""},
    )



@rh_router.get("/ponto/abonos")
async def listar_abonos(
    funcionario_id: Optional[str] = None,
    mes: Optional[int] = None,
    ano: Optional[int] = None,
):
    query = {}
    if funcionario_id:
        query["funcionario_id"] = funcionario_id
    if mes and ano:
        ini = f"{ano}-{mes:02d}-01"
        if mes == 12:
            fim = f"{ano + 1}-01-01"
        else:
            fim = f"{ano}-{mes + 1:02d}-01"
        query["data"] = {"$gte": ini, "$lt": fim}
    abonos = []
    async for a in ponto_abonos_collection.find(query, {"_id": 0}).sort("data", 1):
        abonos.append(a)
    return abonos


@rh_router.delete("/ponto/abono/{abono_id}")
async def deletar_abono(abono_id: str):
    res = await ponto_abonos_collection.delete_one({"id": abono_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Abono não encontrado")
    return {"message": "Abono removido"}


# ===== OBSERVAÇÕES DO MÊS POR FUNCIONÁRIO =====
@rh_router.post("/ponto/observacao")
async def upsert_observacao(payload: dict = Body(...)):
    """Cria ou atualiza observação livre por funcionário/mês/ano.
    Body: { funcionario_id, mes, ano, texto }
    """
    funcionario_id = (payload.get("funcionario_id") or "").strip()
    mes = payload.get("mes")
    ano = payload.get("ano")
    texto = (payload.get("texto") or "").strip()
    
    if not funcionario_id:
        raise HTTPException(status_code=400, detail="funcionario_id é obrigatório")
    if not isinstance(mes, int) or not isinstance(ano, int):
        raise HTTPException(status_code=400, detail="mes e ano (inteiros) são obrigatórios")
    
    chave = {"funcionario_id": funcionario_id, "mes": mes, "ano": ano}
    
    if not texto:
        # Remove se vazio
        await ponto_observacoes_collection.delete_many(chave)
        return {"message": "Observação removida"}
    
    await ponto_observacoes_collection.update_one(
        chave,
        {"$set": {
            **chave,
            "texto": texto,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }, "$setOnInsert": {"id": str(uuid.uuid4())}},
        upsert=True,
    )
    obs = await ponto_observacoes_collection.find_one(chave, {"_id": 0})
    return obs or {"funcionario_id": funcionario_id, "mes": mes, "ano": ano, "texto": texto}


@rh_router.get("/ponto/observacao")
async def get_observacao(funcionario_id: str, mes: int, ano: int):
    obs = await ponto_observacoes_collection.find_one(
        {"funcionario_id": funcionario_id, "mes": mes, "ano": ano},
        {"_id": 0},
    )
    return obs or {"funcionario_id": funcionario_id, "mes": mes, "ano": ano, "texto": ""}







@rh_router.get("/ponto/relatorio-mensal")
async def get_relatorio_ponto_mensal(mes: int, ano: int, funcionario_id: Optional[str] = None):
    """Obter relatório mensal de ponto com cálculo de horas extras e banco de horas"""
    
    JORNADA_SEG_SEX = 8 * 60 + 30
    JORNADA_SABADO = 4 * 60
    
    inicio_mes = f"{ano}-{mes:02d}-01"
    if mes == 12:
        fim_mes = f"{ano + 1}-01-01"
    else:
        fim_mes = f"{ano}-{mes + 1:02d}-01"
    
    query = {"data": {"$gte": inicio_mes, "$lt": fim_mes}}
    if funcionario_id:
        query["funcionario_id"] = funcionario_id
    
    funcionarios_ponto = {}
    async for reg in ponto_collection.find(query):
        fid = reg["funcionario_id"]
        if fid not in funcionarios_ponto:
            funcionarios_ponto[fid] = []
        funcionarios_ponto[fid].append(reg)
    
    relatorio = []
    for fid, registros in funcionarios_ponto.items():
        func = await funcionarios_collection.find_one({"id": fid})
        if not func:
            continue
        
        total_horas_trabalhadas = 0
        total_horas_previstas = 0
        total_atrasos = 0
        dias_trabalhados = 0
        dias_falta = 0
        
        for reg in registros:
            if not reg.get("entrada") or not reg.get("saida"):
                dias_falta += 1
                continue
            
            dias_trabalhados += 1
            
            try:
                data_reg = datetime.strptime(reg["data"], "%Y-%m-%d")
                dia_semana = data_reg.weekday()
            except:
                dia_semana = 0
            
            if dia_semana == 5:
                jornada_prevista = JORNADA_SABADO
            elif dia_semana == 6:
                jornada_prevista = 0
            else:
                jornada_prevista = JORNADA_SEG_SEX
            
            total_horas_previstas += jornada_prevista
            
            try:
                h_ent, m_ent = map(int, reg["entrada"].split(":"))
                h_sai, m_sai = map(int, reg["saida"].split(":"))
                
                minutos_trabalhados = (h_sai * 60 + m_sai) - (h_ent * 60 + m_ent)
                
                if reg.get("saida_almoco") and reg.get("retorno_almoco"):
                    h_sal, m_sal = map(int, reg["saida_almoco"].split(":"))
                    h_ret, m_ret = map(int, reg["retorno_almoco"].split(":"))
                    intervalo = (h_ret * 60 + m_ret) - (h_sal * 60 + m_sal)
                    minutos_trabalhados -= intervalo
                
                total_horas_trabalhadas += minutos_trabalhados
                
                if h_ent > 8 or (h_ent == 8 and m_ent > 0):
                    total_atrasos += (h_ent * 60 + m_ent) - (8 * 60)
            except:
                pass
        
        banco_horas = total_horas_trabalhadas - total_horas_previstas
        horas_extras = max(0, banco_horas)
        horas_devidas = abs(min(0, banco_horas))
        
        salario = func.get("salario", 0)
        valor_hora = salario / 220 if salario > 0 else 0
        valor_hora_extra = valor_hora * 1.5
        valor_horas_extras = (horas_extras / 60) * valor_hora_extra
        
        relatorio.append({
            "funcionario_id": fid,
            "nome": func.get("nome", ""),
            "cargo": func.get("cargo", ""),
            "dias_trabalhados": dias_trabalhados,
            "dias_falta": dias_falta,
            "horas_previstas": f"{total_horas_previstas // 60}h {total_horas_previstas % 60}min",
            "horas_trabalhadas": f"{total_horas_trabalhadas // 60}h {total_horas_trabalhadas % 60}min",
            "horas_extras": f"{horas_extras // 60}h {horas_extras % 60}min",
            "horas_devidas": f"{horas_devidas // 60}h {horas_devidas % 60}min",
            "banco_horas_minutos": banco_horas,
            "total_atrasos_minutos": total_atrasos,
            "total_atrasos": f"{total_atrasos // 60}h {total_atrasos % 60}min",
            "valor_hora": valor_hora,
            "valor_hora_extra": valor_hora_extra,
            "valor_horas_extras": valor_horas_extras
        })
    
    return {
        "mes": mes,
        "ano": ano,
        "funcionarios": relatorio,
        "total_funcionarios": len(relatorio)
    }


# ===== FOLHA DE PAGAMENTO =====
@rh_router.get("/folha-pagamento")
async def list_folha_pagamento(mes: int, ano: int):
    """Listar folhas de pagamento do mês"""
    folhas = []
    async for folha in folha_pagamento_collection.find({"mes": mes, "ano": ano}):
        folha["_id"] = str(folha["_id"])
        func = await funcionarios_collection.find_one({"id": folha["funcionario_id"]})
        folha["funcionario_nome"] = func["nome"] if func else "-"
        folha["funcionario_cargo"] = func.get("cargo", "-") if func else "-"
        folhas.append(folha)
    return folhas


@rh_router.post("/folha-pagamento")
async def create_folha_pagamento(data: FolhaCreate):
    """Criar folha de pagamento"""
    folha_doc = data.dict()
    folha_doc["id"] = str(uuid.uuid4())
    folha_doc["created_at"] = datetime.now().isoformat()
    
    await folha_pagamento_collection.insert_one(folha_doc)
    folha_doc["_id"] = str(folha_doc.get("_id", ""))
    return folha_doc


@rh_router.delete("/folha-pagamento/{folha_id}")
async def delete_folha_pagamento(folha_id: str):
    """Excluir folha de pagamento"""
    result = await folha_pagamento_collection.delete_one({"id": folha_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Folha não encontrada")
    return {"message": "Folha excluída"}


@rh_router.get("/folha-pagamento/{folha_id}/holerite")
async def gerar_holerite(folha_id: str):
    """Gerar PDF do holerite com layout corporativo padronizado."""
    folha = await folha_pagamento_collection.find_one({"id": folha_id})
    if not folha:
        raise HTTPException(status_code=404, detail="Folha não encontrada")
    
    func = await funcionarios_collection.find_one({"id": folha["funcionario_id"]})
    if not func:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    
    pdf_bytes = _build_holerite_pdf(folha, func)
    nome_seguro = (func.get("nome") or "funcionario").replace(" ", "_")[:40]
    filename = f"Holerite_{nome_seguro}_{folha['mes']:02d}_{folha['ano']}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _build_holerite_pdf(folha: dict, func: dict) -> bytes:
    """Gera bytes do PDF de holerite (reutilizável pelo chatbot e rotas REST)."""
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from utils.pdf_template import (
        create_corporate_doc, add_corporate_header, add_footer,
        get_corporate_styles, build_data_table, build_signatures_table,
        BRAND_COLORS, header_table_style,
    )
    
    meses = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    
    def _brl(v):
        try:
            return f"R$ {float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return "R$ 0,00"
    
    buffer = io.BytesIO()
    doc = create_corporate_doc(
        buffer,
        title=f"Holerite {meses[folha['mes']]}/{folha['ano']}",
    )
    styles = get_corporate_styles()
    
    elements = []
    add_corporate_header(
        elements,
        doc_title="HOLERITE DE PAGAMENTO",
        subtitle=f"Competência: {meses[folha['mes']]} / {folha['ano']}",
    )
    
    # Identificação do funcionário
    elements.append(Paragraph("IDENTIFICAÇÃO DO FUNCIONÁRIO", styles["section"]))
    elements.append(build_data_table([
        ("Funcionário:", func.get("nome", "-")),
        ("CPF:", func.get("cpf", "-")),
        ("Cargo:", func.get("cargo", "-")),
        ("Departamento:", func.get("departamento", "-")),
        ("Admissão:", func.get("data_admissao", "-")),
    ]))
    elements.append(Spacer(1, 12))
    
    # Tabela Proventos × Descontos
    horas_extras_total = (folha.get("horas_extras", 0) or 0) * (folha.get("valor_hora_extra", 0) or 0)
    proventos = [
        ("Salário Base", folha.get("salario_base", 0)),
        ("Horas Extras", horas_extras_total),
        ("Adicional Noturno", folha.get("adicional_noturno", 0)),
        ("Comissões", folha.get("comissoes", 0)),
    ]
    descontos = [
        ("INSS", folha.get("inss", 0)),
        ("IRPF", folha.get("irpf", 0)),
        ("Vale Transporte", folha.get("vale_transporte", 0)),
        ("Vale Alimentação", folha.get("vale_alimentacao", 0)),
        ("Plano de Saúde", folha.get("plano_saude", 0)),
        ("Outros Descontos", folha.get("outros_descontos", 0)),
    ]
    
    rows = [["Discriminação", "Proventos", "Descontos"]]
    proventos_filtrados = [(d, v) for d, v in proventos if v > 0]
    descontos_filtrados = [(d, v) for d, v in descontos if v > 0]
    max_len = max(len(proventos_filtrados), len(descontos_filtrados), 1)
    for i in range(max_len):
        if i < len(proventos_filtrados):
            d_p, v_p = proventos_filtrados[i]
            rows.append([d_p, _brl(v_p), ""])
        elif i < len(descontos_filtrados):
            d_d, v_d = descontos_filtrados[i]
            rows.append([d_d, "", _brl(v_d)])
    rows.append([
        "TOTAIS",
        _brl(folha.get("salario_bruto", 0)),
        _brl(folha.get("total_descontos", 0)),
    ])
    
    t_pd = Table(rows, colWidths=[8 * cm, 4 * cm, 4 * cm])
    style = header_table_style()
    style.add("ALIGN", (1, 1), (-1, -1), "RIGHT")
    style.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
    style.add("BACKGROUND", (0, -1), (-1, -1), BRAND_COLORS["label_bg"])
    style.add("LINEABOVE", (0, -1), (-1, -1), 1, BRAND_COLORS["primary"])
    t_pd.setStyle(style)
    elements.append(t_pd)
    elements.append(Spacer(1, 14))
    
    # Total líquido em destaque
    liquido = folha.get("salario_liquido", 0)
    t_liq = Table(
        [["SALÁRIO LÍQUIDO", _brl(liquido)]],
        colWidths=[8 * cm, 8 * cm],
    )
    t_liq.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), BRAND_COLORS["primary"]),
        ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
        ("BACKGROUND", (1, 0), (1, 0), BRAND_COLORS["accent"]),
        ("TEXTCOLOR", (1, 0), (1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 14),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
    ]))
    elements.append(t_liq)
    elements.append(Spacer(1, 14))
    
    # Informações complementares
    elements.append(build_data_table([
        ("Salário Bruto:", _brl(folha.get("salario_bruto", 0))),
        ("Total Descontos:", _brl(folha.get("total_descontos", 0))),
        ("FGTS Depositado:", _brl(folha.get("fgts", 0))),
        ("Base INSS:", _brl(folha.get("base_inss", folha.get("salario_bruto", 0)))),
    ]))
    elements.append(Spacer(1, 24))
    
    # Assinaturas
    elements.append(build_signatures_table(
        left_label="Assinatura do Funcionário",
        right_label="Assinatura do Empregador",
    ))
    
    add_footer(elements, "Sistema CRA · Holerite de Pagamento")
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


@rh_router.post("/folha-pagamento/gerar-contas-pagar")
async def gerar_contas_pagar_folha(mes: int = Body(...), ano: int = Body(...)):
    """Gerar contas a pagar a partir da folha de pagamento"""
    folhas = []
    async for folha in folha_pagamento_collection.find({"mes": mes, "ano": ano}):
        folhas.append(folha)
    
    if not folhas:
        raise HTTPException(status_code=404, detail="Nenhuma folha encontrada para este período")
    
    contas_criadas = []
    
    total_salarios = sum(f.get("salario_liquido", 0) for f in folhas)
    if total_salarios > 0:
        conta_salarios = {
            "id": str(uuid.uuid4()),
            "descricao": f"Folha de Pagamento - {mes:02d}/{ano}",
            "valor": total_salarios,
            "data_vencimento": f"{ano}-{mes:02d}-05",
            "categoria": "Salários",
            "status": "pendente",
            "origem": "rh",
            "created_at": datetime.now().isoformat()
        }
        await db.contas_pagar.insert_one(conta_salarios)
        contas_criadas.append(conta_salarios)
    
    total_inss = sum(f.get("inss", 0) for f in folhas)
    if total_inss > 0:
        conta_inss = {
            "id": str(uuid.uuid4()),
            "descricao": f"INSS - {mes:02d}/{ano}",
            "valor": total_inss,
            "data_vencimento": f"{ano}-{mes:02d}-20",
            "categoria": "Impostos",
            "status": "pendente",
            "origem": "rh",
            "created_at": datetime.now().isoformat()
        }
        await db.contas_pagar.insert_one(conta_inss)
        contas_criadas.append(conta_inss)
    
    total_fgts = sum(f.get("fgts", 0) for f in folhas)
    if total_fgts > 0:
        conta_fgts = {
            "id": str(uuid.uuid4()),
            "descricao": f"FGTS - {mes:02d}/{ano}",
            "valor": total_fgts,
            "data_vencimento": f"{ano}-{mes:02d}-07",
            "categoria": "FGTS",
            "status": "pendente",
            "origem": "rh",
            "created_at": datetime.now().isoformat()
        }
        await db.contas_pagar.insert_one(conta_fgts)
        contas_criadas.append(conta_fgts)
    
    return {
        "message": f"{len(contas_criadas)} contas criadas",
        "contas": contas_criadas
    }


# ===== FÉRIAS =====
@rh_router.get("/ferias")
async def list_ferias(ano: int):
    """Listar férias do ano"""
    ferias_list = []
    async for f in ferias_collection.find({}):
        try:
            inicio = datetime.strptime(f["data_inicio"], "%Y-%m-%d")
            if inicio.year == ano:
                f["_id"] = str(f["_id"])
                ferias_list.append(f)
        except:
            pass
    return ferias_list


@rh_router.get("/ferias/alertas")
async def get_ferias_alertas():
    """Alertas de período aquisitivo"""
    alertas = []
    hoje = datetime.now()

    # Carrega IDs dispensados pelo admin (não exibir nos alertas)
    dispensados_ids = set()
    async for d in db.ferias_alertas_dispensados.find({}, {"_id": 0, "funcionario_id": 1}):
        dispensados_ids.add(d.get("funcionario_id"))

    async for func in funcionarios_collection.find({"status": "ativo"}):
        if func.get("id") in dispensados_ids:
            continue  # admin descartou esse alerta
        if func.get("data_admissao"):
            try:
                data_adm = datetime.strptime(func["data_admissao"], "%Y-%m-%d")
                meses_trabalhados = (hoje.year - data_adm.year) * 12 + (hoje.month - data_adm.month)
                
                if meses_trabalhados >= 11:
                    ultima_ferias = await ferias_collection.find_one(
                        {"funcionario_id": func["id"]},
                        sort=[("data_inicio", -1)]
                    )
                    
                    if not ultima_ferias or (hoje - datetime.strptime(ultima_ferias["data_inicio"], "%Y-%m-%d")).days > 365:
                        alertas.append({
                            "funcionario_id": func["id"],
                            "funcionario_nome": func["nome"],
                            "mensagem": f"Completou {meses_trabalhados} meses - programar férias"
                        })
            except:
                pass
    
    return alertas


@rh_router.post("/ferias")
async def create_ferias(data: FeriasCreate):
    """Agendar férias"""
    ferias_doc = data.dict()
    ferias_doc["id"] = str(uuid.uuid4())
    ferias_doc["created_at"] = datetime.now().isoformat()
    
    await ferias_collection.insert_one(ferias_doc)
    ferias_doc["_id"] = str(ferias_doc.get("_id", ""))
    return ferias_doc


@rh_router.put("/ferias/{ferias_id}")
async def update_ferias(ferias_id: str, data: FeriasCreate):
    """Atualizar férias"""
    existing = await ferias_collection.find_one({"id": ferias_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Férias não encontradas")
    
    update_data = data.dict()
    update_data["updated_at"] = datetime.now().isoformat()
    
    await ferias_collection.update_one({"id": ferias_id}, {"$set": update_data})
    return {"message": "Férias atualizadas com sucesso"}


@rh_router.delete("/ferias/{ferias_id}")
async def delete_ferias(ferias_id: str):
    """Excluir férias"""
    result = await ferias_collection.delete_one({"id": ferias_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Férias não encontradas")
    return {"message": "Férias excluídas com sucesso"}


@rh_router.delete("/ferias")
async def delete_todas_ferias(confirmar: bool = False):
    """Exclui TODOS os registros de férias (zera a coleção).
    Requer query param ?confirmar=true para evitar acidentes."""
    if not confirmar:
        raise HTTPException(
            status_code=400,
            detail="Operação destrutiva. Reenvie com ?confirmar=true para confirmar.",
        )
    total = await ferias_collection.count_documents({})
    result = await ferias_collection.delete_many({})
    return {
        "message": f"Todos os registros de férias foram excluídos ({result.deleted_count} de {total}).",
        "deleted_count": result.deleted_count,
    }


# ===== EPI =====
# ----- CBO: integração com API pública (com fallback local) -----
# Fonte primária: lucassmacedo/cbo-brasil (JSONs públicos com Ocupação, Família, Sinônimo).
# Em caso de falha de rede, cai para o CBO_DATABASE local hardcoded.
_CBO_REMOTE_URLS = {
    "ocupacao": "https://raw.githubusercontent.com/lucassmacedo/cbo-brasil/master/json/CBO2002%20-%20Ocupacao.json",
    "familia": "https://raw.githubusercontent.com/lucassmacedo/cbo-brasil/master/json/CBO2002%20-%20Familia.json",
    "sinonimo": "https://raw.githubusercontent.com/lucassmacedo/cbo-brasil/master/json/CBO2002%20-%20Sinonimo.json",
}
_CBO_REMOTE_TTL = timedelta(hours=24)
_cbo_remote_cache: dict = {
    "loaded_at": None,
    "ocupacoes": [],   # lista de {codigo, ocupacao, familia, sinonimos}
    "available": False,
}
_cbo_remote_lock = threading.Lock()


def _format_cbo_codigo(raw: str) -> str:
    """Converte códigos do dataset (`010105`) para o formato exibido `0101-05`."""
    s = (raw or "").strip().replace("-", "").replace(".", "")
    if len(s) >= 6 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}"
    return raw


async def _carregar_cbo_remoto(force: bool = False) -> bool:
    """Baixa (e cacheia em memória) os dados públicos da CBO.
    Retorna True se o cache está disponível (dados carregados)."""
    now = datetime.now(timezone.utc)
    if (
        not force
        and _cbo_remote_cache["available"]
        and _cbo_remote_cache["loaded_at"]
        and (now - _cbo_remote_cache["loaded_at"]) < _CBO_REMOTE_TTL
    ):
        return True

    try:
        async with httpx.AsyncClient(timeout=15.0, headers={"User-Agent": "CRA-ERP/1.0"}) as client:
            resps = {}
            for key, url in _CBO_REMOTE_URLS.items():
                r = await client.get(url)
                r.raise_for_status()
                resps[key] = r.json()

        # Indexa famílias por código (4 dígitos)
        familias = {item["code"]: item["name"] for item in resps["familia"] if item.get("code")}
        # Sinônimos agrupados por código de ocupação (6 dígitos)
        sinonimos_por_codigo: dict = {}
        for s in resps["sinonimo"]:
            codigo = s.get("code")
            nome = s.get("name")
            if codigo and nome:
                sinonimos_por_codigo.setdefault(codigo, []).append(nome)

        ocupacoes = []
        for o in resps["ocupacao"]:
            codigo = o.get("code") or ""
            nome = o.get("name") or ""
            if not codigo or not nome:
                continue
            familia_codigo = codigo[:4]
            ocupacoes.append({
                "codigo": _format_cbo_codigo(codigo),
                "codigo_raw": codigo,
                "ocupacao": nome,
                "familia": familias.get(familia_codigo, ""),
                "sinonimos": sinonimos_por_codigo.get(codigo, []),
            })

        with _cbo_remote_lock:
            _cbo_remote_cache["ocupacoes"] = ocupacoes
            _cbo_remote_cache["loaded_at"] = now
            _cbo_remote_cache["available"] = bool(ocupacoes)
        logger.info("CBO remoto carregado: %d ocupações.", len(ocupacoes))
        return _cbo_remote_cache["available"]
    except Exception as e:
        logger.warning("Falha ao carregar CBO remoto, usando fallback local: %s", e)
        return False


def _buscar_cbo_remoto(q: str) -> list:
    """Busca em memória usando o cache remoto (já carregado). Não faz I/O."""
    if not _cbo_remote_cache["available"]:
        return []
    q_lower = q.lower().strip()
    q_norm = q_lower.replace("-", "").replace(".", "")
    is_numeric = q_norm.isdigit() and len(q_norm) >= 2

    exatos, parciais_codigo, por_nome = [], [], []
    for item in _cbo_remote_cache["ocupacoes"]:
        codigo_norm = item["codigo_raw"].lower()
        codigo_disp = item["codigo"].lower()

        # Exato por código
        if is_numeric and (q_norm == codigo_norm or q_lower == codigo_disp):
            exatos.append({**item, "match_type": "exact",
                           "descricao": f"Família: {item['familia']}" if item["familia"] else ""})
            continue

        # Parcial por código (numérico)
        if is_numeric and codigo_norm.startswith(q_norm):
            parciais_codigo.append({**item, "match_type": "partial_code",
                                    "descricao": f"Família: {item['familia']}" if item["familia"] else ""})
            continue

        # Nome ou sinônimos
        if not is_numeric and len(q_lower) >= 2:
            if q_lower in item["ocupacao"].lower() or any(q_lower in s.lower() for s in item["sinonimos"]):
                por_nome.append({**item, "match_type": "name",
                                 "descricao": f"Família: {item['familia']}" if item["familia"] else ""})

    if exatos:
        return exatos[:10]
    if parciais_codigo:
        return parciais_codigo[:10]
    return por_nome[:10]


def _buscar_cbo_local(q: str) -> list:
    """Busca usando a base hardcoded local (fallback)."""
    resultados = []
    q_lower = q.lower().strip()
    q_normalized = q_lower.replace("-", "").replace(".", "")

    for codigo, info in CBO_DATABASE.items():
        codigo_normalized = codigo.lower().replace("-", "").replace(".", "")
        if q_normalized == codigo_normalized or q_lower == codigo.lower():
            resultados.append({"codigo": codigo, "ocupacao": info["titulo"],
                               "familia": info["familia"],
                               "descricao": f"Família: {info['familia']}",
                               "match_type": "exact"})
    if resultados:
        return resultados

    for codigo, info in CBO_DATABASE.items():
        codigo_normalized = codigo.lower().replace("-", "").replace(".", "")
        if codigo_normalized.startswith(q_normalized) or codigo.lower().startswith(q_lower):
            resultados.append({"codigo": codigo, "ocupacao": info["titulo"],
                               "familia": info["familia"],
                               "descricao": f"Família: {info['familia']}",
                               "match_type": "partial_code"})
    if resultados:
        return resultados[:10]

    for codigo, info in CBO_DATABASE.items():
        if (q_lower in info["titulo"].lower() or
                any(q_lower in s.lower() for s in info.get("sinonimos", []))):
            resultados.append({"codigo": codigo, "ocupacao": info["titulo"],
                               "familia": info["familia"],
                               "descricao": f"Família: {info['familia']}",
                               "match_type": "name"})
    return resultados[:10]


@rh_router.get("/epi/cbo/buscar")
async def buscar_cbo(q: str, refresh: bool = False):
    """Buscar ocupação por código CBO ou nome.
    Tenta primeiro a API pública (CBO 2002 oficial, ~2600 ocupações + sinônimos).
    Em caso de falha de rede, cai para a base local."""
    q = (q or "").strip()
    if not q:
        return []

    # Tenta carregar/usar o dataset remoto
    remoto_ok = await _carregar_cbo_remoto(force=refresh)
    if remoto_ok:
        resultados = _buscar_cbo_remoto(q)
        if resultados:
            # Remove campos internos antes de devolver
            return [
                {k: v for k, v in r.items() if k != "codigo_raw"}
                for r in resultados
            ]

    # Fallback: base local
    return _buscar_cbo_local(q)


@rh_router.post("/epi/consultar-epis-cbo")
async def consultar_epis_por_cbo(codigo_cbo: str = Body(...), ocupacao: str = Body(...)):
    """Consultar EPIs recomendados por código CBO.
    
    PRIORIZA os documentos normativos da empresa (PGR, PCMSO, LTCAT) carregados na
    Base de Conhecimento. Se a função estiver mapeada nesses documentos, traz EPIs
    EXATOS daqueles laudos. Caso contrário, usa Gemini com NRs gerais.
    """
    
    # Normalizar código CBO
    codigo_cbo = codigo_cbo.strip()

    # Carrega o contexto de PGR/PCMSO/LTCAT/CCT (se disponíveis)
    kb_context = ""
    try:
        from routes.chatbot import _build_knowledge_base_context
        kb_context = await _build_knowledge_base_context()
    except Exception as kb_err:
        logger.warning(f"KB indisponível na consulta de EPIs: {kb_err}")

    # SEMPRE consultar a IA Gemini com o código CBO para garantir precisão
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        kb_prompt_block = ""
        if kb_context:
            kb_prompt_block = (
                "\n\n=== DOCUMENTOS NORMATIVOS DA EMPRESA (USE COMO PRIMEIRA FONTE) ===\n"
                "Os EPIs e riscos abaixo foram levantados no PGR, PCMSO e LTCAT da CRA "
                "Construtora. Quando a função informada estiver coberta por estes documentos, "
                "EXTRAIA os EPIs e riscos DELES (com os números de C.A. e validade quando "
                "houver). Apenas complemente com NRs gerais se a função não estiver mapeada.\n"
                f"{kb_context[:60000]}\n"
                "=== FIM DOS DOCUMENTOS DA EMPRESA ===\n"
            )

        prompt = f"""Você é um especialista em segurança do trabalho no Brasil, com profundo conhecimento da Classificação Brasileira de Ocupações (CBO) e das Normas Regulamentadoras (NRs).{kb_prompt_block}

CÓDIGO CBO INFORMADO: {codigo_cbo}
OCUPAÇÃO: {ocupacao}

IMPORTANTE: Primeiro, verifique se o código CBO {codigo_cbo} corresponde EXATAMENTE à ocupação "{ocupacao}". 
Se houver divergência, informe a ocupação correta para este código CBO.

Liste TODOS os Equipamentos de Proteção Individual (EPIs) obrigatórios e recomendados para a ocupação informada.
PRIORIZE EPIs encontrados nos DOCUMENTOS DA EMPRESA acima (PGR/PCMSO/LTCAT). Se a função estiver
listada nesses documentos, copie nome, C.A. e validade DE LÁ. Apenas para EPIs não cobertos use NRs gerais.

Para cada EPI, forneça:
1. Nome do EPI (igual ao do PGR quando disponível)
2. CA (Certificado de Aprovação) - traga o número exato do PGR/PCMSO se houver, senão "A definir"
3. Validade média em meses (default 12)
4. Prioridade: "Alta" (obrigatório por NR ou listado no PGR como tal), "Média" (recomendado), "Baixa" (opcional)
5. Fonte: "PGR", "PCMSO", "LTCAT", "CCT" ou "NR_geral" — indica de onde a recomendação foi extraída

Também forneça um mapa de risco com os principais riscos da ocupação. Use os riscos do GHE/setor
do PGR quando a função estiver lá; complemente com NRs gerais quando necessário.

Responda APENAS em formato JSON válido:
{{
  "codigo_cbo": "{codigo_cbo}",
  "ocupacao_oficial": "Nome oficial da ocupação segundo CBO",
  "fonte_principal": "PGR" | "PCMSO" | "LTCAT" | "CCT" | "NR_geral",
  "epis": [
    {{"nome": "Nome do EPI", "ca": "12345 ou A definir", "validade_meses": 36, "prioridade": "Alta", "fonte": "PGR"}}
  ],
  "mapa_risco": [
    {{"risco": "Descrição do risco", "prioridade": "Alta", "epi_recomendado": "Nome do EPI", "fonte": "PGR"}}
  ]
}}"""
        
        llm = LlmChat(
            api_key=os.environ.get("EMERGENT_LLM_KEY"),
            session_id=f"epi_cbo_{codigo_cbo}_{datetime.now().strftime('%Y%m%d%H%M%S')}",
            system_message=(
                "Você é especialista em segurança do trabalho, CBO e EPIs no Brasil. "
                "Sempre PRIORIZE os documentos normativos da empresa (PGR, PCMSO, LTCAT) "
                "quando fornecidos. Apenas complemente com NRs gerais se necessário."
            )
        ).with_model("gemini", "gemini-2.5-flash")
        
        response_text = await llm.send_message(UserMessage(text=prompt))
        
        if hasattr(response_text, 'content'):
            response_text = response_text.content
        
        response_text = str(response_text).strip()
        
        # Limpar resposta para extrair JSON
        if response_text.startswith("```"):
            lines = response_text.split("\n")
            json_lines = []
            in_json = False
            for line in lines:
                if line.startswith("```json"):
                    in_json = True
                    continue
                elif line.startswith("```"):
                    in_json = False
                    continue
                elif in_json:
                    json_lines.append(line)
            response_text = "\n".join(json_lines)
        
        result = json.loads(response_text)
        result["fonte"] = "IA_GEMINI"
        result["codigo_consultado"] = codigo_cbo
        return result
        
    except Exception as e:
        # Fallback: usar base local se IA falhar
        categoria = get_categoria_epi(codigo_cbo)
        epis_base = EPI_POR_CARGO.get(categoria, EPI_POR_CARGO["construcao_civil"])
        
        EPI_RISK_MAP = {
            "Capacete": {"risco": "Queda de objetos/impactos na cabeça", "prioridade": "Alta"},
            "Óculos de proteção": {"risco": "Projeção de partículas", "prioridade": "Alta"},
            "Luvas de proteção": {"risco": "Cortes e abrasões", "prioridade": "Alta"},
            "Botina de segurança": {"risco": "Queda de objetos nos pés", "prioridade": "Alta"},
            "Protetor auricular": {"risco": "Ruído excessivo", "prioridade": "Média"},
            "Cinto de segurança": {"risco": "Queda de altura", "prioridade": "Alta"},
            "Capacete classe B": {"risco": "Choque elétrico e impactos", "prioridade": "Alta"},
            "Luvas isolantes": {"risco": "Choque elétrico", "prioridade": "Alta"},
            "Botina isolante": {"risco": "Choque elétrico pelos pés", "prioridade": "Alta"},
            "Vestimenta antichama": {"risco": "Arco elétrico/chamas", "prioridade": "Alta"},
            "Detector de tensão": {"risco": "Trabalho com energia viva", "prioridade": "Alta"},
            "Máscara de solda": {"risco": "Radiação e respingos", "prioridade": "Alta"},
            "Avental de raspa": {"risco": "Respingos de solda", "prioridade": "Alta"},
            "Luvas de raspa": {"risco": "Queimaduras nas mãos", "prioridade": "Alta"},
        }
        
        epis = []
        mapa_risco = []
        
        for epi in epis_base:
            epis.append({
                "nome": epi["nome"],
                "ca": epi.get("ca", "A definir"),
                "validade_meses": epi.get("validade_meses", 12),
                "prioridade": epi.get("prioridade", "Alta")
            })
            
            if epi["nome"] in EPI_RISK_MAP:
                risk_info = EPI_RISK_MAP[epi["nome"]]
                mapa_risco.append({
                    "risco": risk_info["risco"],
                    "prioridade": risk_info["prioridade"],
                    "epi_recomendado": epi["nome"]
                })
        
        return {
            "codigo_cbo": codigo_cbo,
            "ocupacao_oficial": ocupacao,
            "epis": epis,
            "mapa_risco": mapa_risco,
            "fonte": "FALLBACK_LOCAL",
            "erro": str(e)
        }


@rh_router.get("/epi/fichas")
async def list_fichas_epi(funcionario_id: Optional[str] = None):
    """Listar fichas de EPI"""
    query = {}
    if funcionario_id:
        query["funcionario_id"] = funcionario_id
    
    fichas = []
    async for ficha in epi_fichas_collection.find(query, {"_id": 0}):
        func = await funcionarios_collection.find_one({"id": ficha["funcionario_id"]}, {"_id": 0})
        ficha["funcionario_nome"] = func["nome"] if func else "-"
        fichas.append(ficha)
    
    return fichas


@rh_router.post("/epi/fichas")
async def create_ficha_epi(data: FichaEPICreate):
    """Criar ficha de EPI"""
    ficha_doc = data.dict()
    ficha_doc["id"] = str(uuid.uuid4())
    ficha_doc["created_at"] = datetime.now().isoformat()

    # Auto-preenche cargo a partir do funcionário se não foi enviado
    if not ficha_doc.get("cargo"):
        func = await funcionarios_collection.find_one(
            {"id": ficha_doc["funcionario_id"]}, {"_id": 0}
        )
        ficha_doc["cargo"] = (func or {}).get("cargo") or ficha_doc.get("ocupacao_cbo") or "-"

    await epi_fichas_collection.insert_one(ficha_doc)
    # Remove o _id que o Mongo adicionou para evitar erro de serialização
    ficha_doc.pop("_id", None)
    return ficha_doc


def _build_ficha_epi_pdf(ficha: dict, func: dict) -> bytes:
    """Gera o PDF da Ficha de EPI completa com layout corporativo padronizado."""
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, KeepTogether
    from utils.pdf_template import (
        create_corporate_doc, add_corporate_header, add_footer,
        get_corporate_styles, build_data_table, build_signatures_table,
        BRAND_COLORS, header_table_style,
    )

    def _br_date(s: str) -> str:
        if not s:
            return "-"
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            return s

    buffer = io.BytesIO()
    doc = create_corporate_doc(
        buffer,
        title=f"Ficha de EPI - {func.get('nome','funcionario')}",
    )
    styles = get_corporate_styles()

    elements = []
    add_corporate_header(
        elements,
        doc_title="FICHA DE CONTROLE DE EPI",
        subtitle="Equipamento de Proteção Individual — NR-06",
    )

    # Dados do funcionário
    elements.append(Paragraph("IDENTIFICAÇÃO DO FUNCIONÁRIO", styles["section"]))
    elements.append(build_data_table([
        ("Funcionário:", func.get("nome", "-")),
        ("CPF:", func.get("cpf", "-")),
        ("RG:", func.get("rg", "-")),
        ("Cargo:", ficha.get("cargo") or func.get("cargo", "-")),
        ("CBO:", f"{ficha.get('codigo_cbo','-')} - {ficha.get('ocupacao_cbo','')}".rstrip(" -")),
        ("Departamento:", func.get("departamento", "-")),
        ("Admissão:", _br_date(func.get("data_admissao", ""))),
        ("Data de Entrega:", _br_date(ficha.get("data_entrega", ""))),
    ]))
    elements.append(Spacer(1, 14))

    # Tabela de EPIs entregues
    elements.append(Paragraph("EQUIPAMENTOS DE PROTEÇÃO INDIVIDUAL ENTREGUES", styles["section"]))
    rows = [["#", "EPI", "C.A.", "Validade", "Prioridade"]]
    epis = ficha.get("epis") or []
    for idx, e in enumerate(epis, 1):
        rows.append([
            str(idx),
            (e.get("nome") or "-"),
            (e.get("ca") or "-"),
            _br_date(e.get("validade") or ""),
            (e.get("prioridade") or "Média"),
        ])
    if not epis:
        rows.append(["-", "Nenhum EPI registrado", "-", "-", "-"])

    t_epis = Table(
        rows,
        colWidths=[0.8 * cm, 8.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm],
        repeatRows=1,
    )
    style = header_table_style()
    style.add("ALIGN", (0, 1), (0, -1), "CENTER")
    style.add("ALIGN", (2, 1), (-1, -1), "CENTER")
    t_epis.setStyle(style)
    elements.append(t_epis)
    elements.append(Spacer(1, 14))

    # Observações
    obs = (ficha.get("observacoes") or "").strip()
    if obs:
        elements.append(Paragraph("OBSERVAÇÕES", styles["section"]))
        elements.append(Paragraph(obs.replace("\n", "<br/>"), styles["body"]))
        elements.append(Spacer(1, 12))

    # Termo de recebimento (texto compacto)
    elements.append(Paragraph("TERMO DE RECEBIMENTO E RESPONSABILIDADE", styles["section"]))
    termo_txt = (
        "Declaro, para os devidos fins, que recebi gratuitamente da empresa CRA Construtora "
        "os Equipamentos de Proteção Individual (EPI) discriminados nesta ficha, em perfeitas "
        "condições de uso, comprometendo-me a: <br/>"
        "<b>I –</b> usá-los apenas para a finalidade a que se destinam, durante toda a jornada de trabalho;<br/>"
        "<b>II –</b> ser responsável pela guarda e conservação;<br/>"
        "<b>III –</b> comunicar imediatamente ao empregador qualquer alteração que torne o EPI "
        "impróprio para uso ou que necessite de reposição;<br/>"
        "<b>IV –</b> cumprir as determinações da empresa quanto ao uso adequado;<br/>"
        "<b>V –</b> devolvê-los ao empregador quando não for mais utilizá-los, em caso de "
        "desligamento ou troca, sob pena das sanções previstas na NR-06 da Portaria 3.214/78."
    )
    elements.append(Paragraph(termo_txt, styles["body"]))
    elements.append(Spacer(1, 30))

    # Assinaturas
    elements.append(KeepTogether(build_signatures_table(
        left_label="Assinatura do Funcionário",
        right_label="Assinatura do Empregador / SESMT",
    )))

    add_footer(elements, "Sistema CRA · Ficha de Controle de EPI - NR-06")

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def _build_termo_responsabilidade_pdf(ficha: dict, func: dict) -> bytes:
    """Gera o Termo de Responsabilidade isolado (NR-06) com layout corporativo."""
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from utils.pdf_template import (
        create_corporate_doc, add_corporate_header, add_footer,
        get_corporate_styles, build_data_table, build_signatures_table,
        BRAND_COLORS,
    )

    def _br_date(s: str) -> str:
        if not s:
            return "-"
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            return s

    buffer = io.BytesIO()
    doc = create_corporate_doc(
        buffer,
        title=f"Termo de Responsabilidade - {func.get('nome','funcionario')}",
    )
    styles = get_corporate_styles()

    elements = []
    add_corporate_header(
        elements,
        doc_title="TERMO DE RESPONSABILIDADE - EPI",
        subtitle="NR-06 — Portaria 3.214/78 do Ministério do Trabalho",
    )

    elements.append(Paragraph("IDENTIFICAÇÃO", styles["section"]))
    elements.append(build_data_table([
        ("Funcionário:", func.get("nome", "-")),
        ("CPF:", func.get("cpf", "-")),
        ("Cargo:", ficha.get("cargo") or func.get("cargo", "-")),
        ("CBO:", f"{ficha.get('codigo_cbo','-')} - {ficha.get('ocupacao_cbo','')}".rstrip(" -")),
        ("Data de Entrega:", _br_date(ficha.get("data_entrega", ""))),
    ]))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("DECLARAÇÃO", styles["section"]))
    declaracao = (
        f"Eu, <b>{func.get('nome','-')}</b>, portador(a) do CPF nº "
        f"<b>{func.get('cpf','-')}</b>, declaro para os devidos fins que recebi "
        f"da <b>CRA Construtora</b>, em "
        f"<b>{_br_date(ficha.get('data_entrega',''))}</b>, gratuitamente e em perfeitas "
        "condições de uso, os Equipamentos de Proteção Individual (EPI) discriminados a "
        "seguir, comprometendo-me a:<br/><br/>"
        "<b>I.</b> Utilizá-los apenas para a finalidade a que se destinam, durante toda "
        "a jornada de trabalho;<br/>"
        "<b>II.</b> Responsabilizar-me pela guarda e conservação dos referidos EPIs;<br/>"
        "<b>III.</b> Comunicar imediatamente ao empregador qualquer alteração que torne "
        "o EPI impróprio para uso ou que necessite de substituição;<br/>"
        "<b>IV.</b> Cumprir as determinações do empregador quanto ao uso adequado;<br/>"
        "<b>V.</b> Devolvê-los à empresa em caso de desligamento, transferência ou "
        "substituição;<br/>"
        "<b>VI.</b> Estar ciente de que o não uso ou uso inadequado dos EPIs constitui "
        "ato faltoso, podendo gerar advertência, suspensão ou demissão por justa causa, "
        "conforme art. 158, parágrafo único, alínea 'b', da CLT e item 6.7.1 da NR-06."
    )
    elements.append(Paragraph(declaracao, styles["body"]))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("EQUIPAMENTOS RECEBIDOS", styles["section"]))
    rows = [["#", "EPI", "C.A.", "Validade"]]
    epis = ficha.get("epis") or []
    for idx, e in enumerate(epis, 1):
        rows.append([
            str(idx),
            (e.get("nome") or "-"),
            (e.get("ca") or "-"),
            _br_date(e.get("validade") or ""),
        ])
    if not epis:
        rows.append(["-", "Nenhum EPI registrado", "-", "-"])

    from utils.pdf_template import header_table_style
    t_epis = Table(rows, colWidths=[0.8 * cm, 11.5 * cm, 2.5 * cm, 2.5 * cm], repeatRows=1)
    style = header_table_style()
    style.add("ALIGN", (0, 1), (0, -1), "CENTER")
    style.add("ALIGN", (2, 1), (-1, -1), "CENTER")
    t_epis.setStyle(style)
    elements.append(t_epis)
    elements.append(Spacer(1, 28))

    elements.append(build_signatures_table(
        left_label="Assinatura do Funcionário",
        right_label="Assinatura do Empregador / SESMT",
    ))

    add_footer(elements, "Sistema CRA · Termo de Responsabilidade NR-06")

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


@rh_router.get("/epi/fichas/{ficha_id}/exportar")
async def exportar_ficha_epi(ficha_id: str):
    """Exportar a Ficha de EPI completa em PDF com layout corporativo."""
    ficha = await epi_fichas_collection.find_one({"id": ficha_id}, {"_id": 0})
    if not ficha:
        raise HTTPException(status_code=404, detail="Ficha de EPI não encontrada")
    func = await funcionarios_collection.find_one({"id": ficha["funcionario_id"]}, {"_id": 0})
    if not func:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    pdf_bytes = _build_ficha_epi_pdf(ficha, func)
    nome_safe = (func.get("nome") or "func").replace(" ", "_")[:40]
    filename = f"FichaEPI_{nome_safe}_{(ficha.get('data_entrega') or '')[:10]}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@rh_router.get("/epi/fichas/{ficha_id}/termo-responsabilidade")
async def exportar_termo_responsabilidade(ficha_id: str):
    """Exportar o Termo de Responsabilidade NR-06 em PDF com layout corporativo."""
    ficha = await epi_fichas_collection.find_one({"id": ficha_id}, {"_id": 0})
    if not ficha:
        raise HTTPException(status_code=404, detail="Ficha de EPI não encontrada")
    func = await funcionarios_collection.find_one({"id": ficha["funcionario_id"]}, {"_id": 0})
    if not func:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    pdf_bytes = _build_termo_responsabilidade_pdf(ficha, func)
    nome_safe = (func.get("nome") or "func").replace(" ", "_")[:40]
    filename = f"TermoResponsabilidade_{nome_safe}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ===== NOTIFICAÇÕES =====
@rh_router.get("/notificacoes")
async def get_rh_notificacoes():
    """Obter todas as notificações do RH"""
    hoje = datetime.now()
    mes_atual = hoje.month

    # Carrega o set de notificações RH dispensadas (uma única consulta)
    # Estrutura: {escopo:"rh", tipo, ref_id}
    dispensados_docs = await db.notificacoes_dispensadas.find(
        {"escopo": "rh"}, {"_id": 0, "tipo": 1, "ref_id": 1}
    ).to_list(2000)
    dispensados = {(d.get("tipo"), d.get("ref_id")) for d in dispensados_docs}

    def _disp(tipo: str, ref_id: str) -> bool:
        return (tipo, ref_id) in dispensados

    aniversariantes = []
    async for func in funcionarios_collection.find({"status": "ativo"}):
        if func.get("data_nascimento"):
            try:
                data_nasc = datetime.strptime(func["data_nascimento"], "%Y-%m-%d")
                if data_nasc.month == mes_atual:
                    # ref_id inclui o mês para que ao virar o mês a dispensa reseta
                    ref = f"{func['id']}::{hoje.year}-{mes_atual:02d}"
                    if _disp("aniversariante", ref):
                        continue
                    idade = hoje.year - data_nasc.year
                    aniversariantes.append({
                        "ref_id": ref,
                        "funcionario_id": func["id"],
                        "nome": func["nome"],
                        "cargo": func.get("cargo", "-"),
                        "data_formatada": data_nasc.strftime("%d/%m"),
                        "idade": idade
                    })
            except:
                pass
    
    alertas_ferias = []
    funcionarios_sem_ferias = []
    async for func in funcionarios_collection.find({"status": "ativo"}):
        if func.get("data_admissao"):
            try:
                data_adm = datetime.strptime(func["data_admissao"], "%Y-%m-%d")
                meses_trabalhados = (hoje.year - data_adm.year) * 12 + (hoje.month - data_adm.month)
                
                ultima_ferias = await ferias_collection.find_one(
                    {"funcionario_id": func["id"]},
                    sort=[("data_inicio", -1)]
                )
                
                if meses_trabalhados >= 11 and meses_trabalhados <= 14:
                    if not _disp("alerta_ferias", func["id"]):
                        alertas_ferias.append({
                            "ref_id": func["id"],
                            "funcionario_id": func["id"],
                            "nome": func["nome"],
                            "mensagem": f"Período aquisitivo completando"
                        })
                
                if meses_trabalhados >= 12:
                    if not ultima_ferias:
                        if not _disp("funcionario_sem_ferias", func["id"]):
                            funcionarios_sem_ferias.append({
                                "ref_id": func["id"],
                                "funcionario_id": func["id"],
                                "nome": func["nome"],
                                "ultima_ferias": "Nunca tirou férias"
                            })
                    else:
                        ultima = datetime.strptime(ultima_ferias["data_inicio"], "%Y-%m-%d")
                        if (hoje - ultima).days > 365:
                            if not _disp("funcionario_sem_ferias", func["id"]):
                                funcionarios_sem_ferias.append({
                                    "ref_id": func["id"],
                                    "funcionario_id": func["id"],
                                    "nome": func["nome"],
                                    "ultima_ferias": ultima.strftime("%d/%m/%Y")
                                })
            except:
                pass
    
    alertas_epi = []
    async for ficha in epi_fichas_collection.find({}):
        func = await funcionarios_collection.find_one({"id": ficha.get("funcionario_id")})
        if func and ficha.get("epis"):
            for epi in ficha["epis"]:
                if epi.get("validade"):
                    try:
                        validade = datetime.strptime(epi["validade"], "%Y-%m-%d")
                        dias_restantes = (validade - hoje).days
                        if 0 < dias_restantes <= 30:
                            ref = f"{func['id']}::{epi.get('nome','')}::{epi['validade']}"
                            if _disp("alerta_epi", ref):
                                continue
                            alertas_epi.append({
                                "ref_id": ref,
                                "funcionario_id": func["id"],
                                "funcionario": func["nome"],
                                "epi": epi["nome"],
                                "dias_restantes": dias_restantes
                            })
                    except:
                        pass
    
    inconsistencias_ponto = []
    hoje_str = hoje.strftime("%Y-%m-%d")
    dia_semana = hoje.weekday()
    
    if dia_semana < 6:
        async for reg in ponto_collection.find({"data": hoje_str}):
            func = await funcionarios_collection.find_one({"id": reg.get("funcionario_id")})
            if func and reg.get("entrada"):
                try:
                    h, m = map(int, reg["entrada"].split(":"))
                    hora_limite = 8
                    minuto_limite = 15
                    
                    if h > hora_limite or (h == hora_limite and m > minuto_limite):
                        ref = f"{func['id']}::{hoje_str}"
                        if _disp("inconsistencia_ponto", ref):
                            continue
                        atraso = (h - 8) * 60 + m
                        inconsistencias_ponto.append({
                            "ref_id": ref,
                            "funcionario_id": func["id"],
                            "funcionario": func["nome"],
                            "tipo": "Atraso",
                            "detalhe": f"Entrada às {reg['entrada']} ({atraso} min de atraso)"
                        })
                except:
                    pass
    
    return {
        "aniversariantes": aniversariantes,
        "alertas_ferias": alertas_ferias,
        "funcionarios_sem_ferias": funcionarios_sem_ferias,
        "alertas_epi": alertas_epi,
        "alertas_atestados": [],
        "inconsistencias_ponto": inconsistencias_ponto
    }


# ===== Dispensar / restaurar notificações RH individualmente =====

class DispensarNotificacaoRH(BaseModel):
    tipo: str  # aniversariante | alerta_ferias | funcionario_sem_ferias | alerta_epi | inconsistencia_ponto
    ref_id: str


@rh_router.post("/notificacoes/dispensar")
async def dispensar_notificacao_rh(payload: DispensarNotificacaoRH):
    """Dispensa (oculta) uma notificação RH específica."""
    tipos_validos = {
        "aniversariante", "alerta_ferias", "funcionario_sem_ferias",
        "alerta_epi", "inconsistencia_ponto",
    }
    if payload.tipo not in tipos_validos:
        raise HTTPException(status_code=400, detail=f"Tipo inválido. Use um de: {tipos_validos}")
    if not payload.ref_id:
        raise HTTPException(status_code=400, detail="ref_id é obrigatório")
    await db.notificacoes_dispensadas.update_one(
        {"escopo": "rh", "tipo": payload.tipo, "ref_id": payload.ref_id},
        {"$set": {
            "escopo": "rh",
            "tipo": payload.tipo,
            "ref_id": payload.ref_id,
            "dispensada_em": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    return {"dispensada": True, "tipo": payload.tipo, "ref_id": payload.ref_id}


@rh_router.delete("/notificacoes/dispensar")
async def restaurar_notificacao_rh(tipo: str, ref_id: str):
    """Restaura uma notificação RH previamente dispensada."""
    res = await db.notificacoes_dispensadas.delete_one(
        {"escopo": "rh", "tipo": tipo, "ref_id": ref_id}
    )
    return {"restaurada": True, "removidos": res.deleted_count}


@rh_router.delete("/notificacoes/dispensar-todos")
async def restaurar_todas_rh():
    """Restaura TODAS as notificações RH dispensadas."""
    res = await db.notificacoes_dispensadas.delete_many({"escopo": "rh"})
    return {"restauradas": res.deleted_count}


@rh_router.get("/notificacoes/contagem")
async def get_rh_notificacoes_contagem():
    """Contagem de notificações para badge"""
    notifs = await get_rh_notificacoes()
    
    total = (
        len(notifs["alertas_ferias"]) +
        len(notifs["funcionarios_sem_ferias"]) +
        len(notifs["alertas_epi"]) +
        len(notifs["inconsistencias_ponto"])
    )
    
    urgentes = len(notifs["funcionarios_sem_ferias"]) + len([e for e in notifs["alertas_epi"] if e.get("dias_restantes", 30) <= 7])
    
    return {"total": total, "urgentes": urgentes}


# ===== CUSTOS =====
@rh_router.get("/custos/config")
async def get_custos_config():
    """Retorna a configuração de custos (cria com defaults se não existir)."""
    cfg = await custos_rh_config_collection.find_one({"id": "default"}, {"_id": 0})
    if cfg:
        return cfg
    defaults = {
        "id": "default",
        "fgts_aliquota": FGTS_ALIQUOTA,
        "fgts_funcionario_ids": [],
        "inss_patronal_aliquota": INSS_PATRONAL_ALIQUOTA,
        "inss_patronal_funcionario_ids": [],
        "vale_transporte": 0.0,
        "vale_transporte_funcionario_ids": [],
        "vale_alimentacao": 0.0,
        "vale_alimentacao_funcionario_ids": [],
        "plano_saude": 0.0,
        "plano_saude_funcionario_ids": [],
        "outros_beneficios": 150.0,
        "outros_beneficios_funcionario_ids": [],
        "epis_custo_mensal": 50.0,
        "epis_custo_mensal_funcionario_ids": [],
        "horas_mes": 220,
        "custos_extras": [],
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await custos_rh_config_collection.insert_one(defaults)
    defaults.pop("_id", None)
    return defaults


@rh_router.put("/custos/config")
async def update_custos_config(payload: dict = Body(...)):
    """Atualiza a configuração global de custos."""
    update = {}
    campos_float = [
        "fgts_aliquota", "inss_patronal_aliquota",
        "vale_transporte", "vale_alimentacao", "plano_saude",
        "outros_beneficios", "epis_custo_mensal",
    ]
    for k in campos_float:
        if k in payload:
            try:
                update[k] = float(payload[k] or 0)
            except (ValueError, TypeError):
                raise HTTPException(status_code=400, detail=f"{k} deve ser numérico")
    if "horas_mes" in payload:
        try:
            update["horas_mes"] = int(payload["horas_mes"]) or 220
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="horas_mes deve ser inteiro")
    
    # Listas de funcionário_ids paralelas a cada campo padrão (vazia = todos)
    campos_lista = [
        "fgts_funcionario_ids", "inss_patronal_funcionario_ids",
        "vale_transporte_funcionario_ids", "vale_alimentacao_funcionario_ids",
        "plano_saude_funcionario_ids", "outros_beneficios_funcionario_ids",
        "epis_custo_mensal_funcionario_ids",
    ]
    for k in campos_lista:
        if k in payload:
            v = payload.get(k) or []
            if not isinstance(v, list):
                raise HTTPException(status_code=400, detail=f"{k} deve ser uma lista")
            update[k] = [str(x) for x in v]
    
    # Custos extras: lista de { id, nome, tipo (fixo|percentual), valor, funcionario_ids: [] }
    if "custos_extras" in payload:
        extras_in = payload.get("custos_extras") or []
        if not isinstance(extras_in, list):
            raise HTTPException(status_code=400, detail="custos_extras deve ser uma lista")
        extras_norm = []
        for it in extras_in:
            if not isinstance(it, dict):
                continue
            nome = (it.get("nome") or "").strip()
            tipo = (it.get("tipo") or "fixo").strip().lower()
            if tipo not in ("fixo", "percentual"):
                tipo = "fixo"
            try:
                valor = float(it.get("valor") or 0)
            except (ValueError, TypeError):
                valor = 0.0
            func_ids = it.get("funcionario_ids") or []
            if not isinstance(func_ids, list):
                func_ids = []
            if not nome:
                continue  # ignora itens sem nome
            extras_norm.append({
                "id": it.get("id") or str(uuid.uuid4()),
                "nome": nome,
                "tipo": tipo,
                "valor": valor,
                "funcionario_ids": [str(x) for x in func_ids],
            })
        update["custos_extras"] = extras_norm
    
    if not update:
        raise HTTPException(status_code=400, detail="Nada para atualizar")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await custos_rh_config_collection.update_one(
        {"id": "default"},
        {"$set": update, "$setOnInsert": {"id": "default"}},
        upsert=True,
    )
    cfg = await custos_rh_config_collection.find_one({"id": "default"}, {"_id": 0})
    return cfg


async def _carregar_custos_config():
    """Carrega config (chama get_custos_config para garantir defaults)."""
    return await get_custos_config()


@rh_router.get("/custos")
async def get_custos_rh():
    """Obter custos detalhados de RH (usa configuração editável)."""
    cfg = await _carregar_custos_config()
    fgts_aliq = float(cfg.get("fgts_aliquota") or FGTS_ALIQUOTA)
    inss_aliq = float(cfg.get("inss_patronal_aliquota") or INSS_PATRONAL_ALIQUOTA)
    vt = float(cfg.get("vale_transporte") or 0)
    va = float(cfg.get("vale_alimentacao") or 0)
    ps = float(cfg.get("plano_saude") or 0)
    outros = float(cfg.get("outros_beneficios") or 0)
    epis_padrao = float(cfg.get("epis_custo_mensal") or 0)
    horas_mes = int(cfg.get("horas_mes") or 220)
    custos_extras = cfg.get("custos_extras") or []
    
    # Listas de funcionário_ids para cada campo padrão (vazio = aplica a todos)
    fgts_fids = set(cfg.get("fgts_funcionario_ids") or [])
    inss_fids = set(cfg.get("inss_patronal_funcionario_ids") or [])
    vt_fids = set(cfg.get("vale_transporte_funcionario_ids") or [])
    va_fids = set(cfg.get("vale_alimentacao_funcionario_ids") or [])
    ps_fids = set(cfg.get("plano_saude_funcionario_ids") or [])
    outros_fids = set(cfg.get("outros_beneficios_funcionario_ids") or [])
    epis_fids = set(cfg.get("epis_custo_mensal_funcionario_ids") or [])
    
    def _aplica(fid: str, lista: set) -> bool:
        return (not lista) or (fid in lista)
    
    custos_funcionarios = []
    total_salarios = 0
    total_encargos = 0
    total_beneficios = 0
    total_epis = 0
    total_extras = 0
    
    async for func in funcionarios_collection.find({"status": "ativo"}):
        salario = float(func.get("salario", 0) or 0)
        fid = func["id"]
        fgts = (salario * (fgts_aliq / 100)) if _aplica(fid, fgts_fids) else 0.0
        inss_patronal = (salario * (inss_aliq / 100)) if _aplica(fid, inss_fids) else 0.0
        vt_f = vt if _aplica(fid, vt_fids) else 0.0
        va_f = va if _aplica(fid, va_fids) else 0.0
        ps_f = ps if _aplica(fid, ps_fids) else 0.0
        outros_f = outros if _aplica(fid, outros_fids) else 0.0
        beneficios = vt_f + va_f + ps_f + outros_f
        epis_custo = epis_padrao if _aplica(fid, epis_fids) else 0.0
        
        # Aplicar custos extras: se funcionario_ids vazio = aplica a todos; senão só aos listados
        extras_funcionario = []
        extras_total = 0.0
        for ce in custos_extras:
            ce_func_ids = ce.get("funcionario_ids") or []
            if ce_func_ids and fid not in ce_func_ids:
                continue
            valor = float(ce.get("valor") or 0)
            if ce.get("tipo") == "percentual":
                aplicado = salario * (valor / 100)
            else:
                aplicado = valor
            extras_total += aplicado
            extras_funcionario.append({
                "nome": ce.get("nome"),
                "tipo": ce.get("tipo"),
                "valor": valor,
                "valor_aplicado": aplicado,
            })
        
        custo_total = salario + fgts + inss_patronal + beneficios + epis_custo + extras_total
        custo_hora = custo_total / horas_mes if custo_total > 0 and horas_mes > 0 else 0
        
        total_salarios += salario
        total_encargos += fgts + inss_patronal
        total_beneficios += beneficios
        total_epis += epis_custo
        total_extras += extras_total
        
        custos_funcionarios.append({
            "funcionario_id": fid,
            "nome": func["nome"],
            "cargo": func.get("cargo", "-"),
            "salario": salario,
            "fgts": fgts,
            "inss_patronal": inss_patronal,
            "beneficios": beneficios,
            "epis": epis_custo,
            "extras_total": extras_total,
            "extras_detalhe": extras_funcionario,
            "custo_total": custo_total,
            "custo_hora": custo_hora,
        })
    
    return {
        "funcionarios": custos_funcionarios,
        "resumo": {
            "total_salarios": total_salarios,
            "total_encargos": total_encargos,
            "total_beneficios": total_beneficios,
            "total_epis": total_epis,
            "total_extras": total_extras,
            "custo_total": total_salarios + total_encargos + total_beneficios + total_epis + total_extras,
        },
        "config": {
            "fgts_aliquota": fgts_aliq,
            "fgts_funcionario_ids": list(fgts_fids),
            "inss_patronal_aliquota": inss_aliq,
            "inss_patronal_funcionario_ids": list(inss_fids),
            "vale_transporte": vt,
            "vale_transporte_funcionario_ids": list(vt_fids),
            "vale_alimentacao": va,
            "vale_alimentacao_funcionario_ids": list(va_fids),
            "plano_saude": ps,
            "plano_saude_funcionario_ids": list(ps_fids),
            "outros_beneficios": outros,
            "outros_beneficios_funcionario_ids": list(outros_fids),
            "epis_custo_mensal": epis_padrao,
            "epis_custo_mensal_funcionario_ids": list(epis_fids),
            "horas_mes": horas_mes,
            "custos_extras": custos_extras,
        },
    }


@rh_router.post("/custos/simular-dissidio")
async def simular_dissidio(percentual: float = Body(..., embed=True)):
    """Simular impacto de dissídio na folha"""
    cfg = await _carregar_custos_config()
    fgts_aliq = float(cfg.get("fgts_aliquota") or FGTS_ALIQUOTA)
    inss_aliq = float(cfg.get("inss_patronal_aliquota") or INSS_PATRONAL_ALIQUOTA)
    folha_atual = 0
    async for func in funcionarios_collection.find({"status": "ativo"}):
        salario = func.get("salario", 0)
        fgts = salario * (fgts_aliq / 100)
        inss_patronal = salario * (inss_aliq / 100)
        folha_atual += salario + fgts + inss_patronal
    
    folha_nova = folha_atual * (1 + percentual / 100)
    impacto_mensal = folha_nova - folha_atual
    
    return {
        "percentual": percentual,
        "folha_atual": folha_atual,
        "folha_nova": folha_nova,
        "impacto_mensal": impacto_mensal,
        "impacto_anual": impacto_mensal * 12
    }


@rh_router.post("/custos/simular-rescisao")
async def simular_rescisao(funcionario_id: str = Body(..., embed=True)):
    """Simular provisão de rescisão"""
    func = await funcionarios_collection.find_one({"id": funcionario_id})
    if not func:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    
    cfg = await _carregar_custos_config()
    fgts_aliq = float(cfg.get("fgts_aliquota") or FGTS_ALIQUOTA)
    
    salario = func.get("salario", 0)
    data_admissao = func.get("data_admissao", "")
    
    try:
        data_adm = datetime.strptime(data_admissao, "%Y-%m-%d")
        meses_trabalhados = (datetime.now().year - data_adm.year) * 12 + (datetime.now().month - data_adm.month)
    except:
        meses_trabalhados = 12
    
    saldo_salario = salario
    aviso_previo = salario + (salario / 30 * 3 * (meses_trabalhados // 12))
    ferias_proporcionais = (salario / 12) * (meses_trabalhados % 12) + (salario / 12 * (meses_trabalhados % 12) / 3)
    decimo_terceiro = (salario / 12) * (datetime.now().month)
    fgts_saldo = salario * (fgts_aliq / 100) * meses_trabalhados
    multa_fgts = fgts_saldo * 0.4
    
    total = saldo_salario + aviso_previo + ferias_proporcionais + decimo_terceiro + multa_fgts
    
    return {
        "funcionario": func["nome"],
        "meses_trabalhados": meses_trabalhados,
        "saldo_salario": saldo_salario,
        "aviso_previo": aviso_previo,
        "ferias_proporcionais": ferias_proporcionais,
        "decimo_terceiro": decimo_terceiro,
        "fgts_saldo": fgts_saldo,
        "multa_fgts": multa_fgts,
        "total": total
    }
