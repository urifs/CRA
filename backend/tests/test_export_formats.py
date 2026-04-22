"""
Test Export Formats - PDF and Excel
Verifies:
1. PDF contas_pagar: columns Descrição, Valor (R$ X,XX), Vencimento (DD/MM/AAAA), Quitação, Status, Fornecedor
2. PDF contas_receber: columns Descrição, Valor (R$ X,XX), Vencimento (DD/MM/AAAA), Recebimento, Status, Cliente
3. Excel contas_pagar: columns Descrição, Valor, Vencimento (DD/MM/AAAA), Quitação, Status, Fornecedor, Centro de Custo, Plano de Contas
4. Excel contas_receber: columns Descrição, Valor, Vencimento (DD/MM/AAAA), Recebimento, Status, Cliente, Centro de Custo, Plano de Contas
5. Excel obras: columns Nome, Local, Cliente, Status, Data Início, Data Fim, Contrato
6. Excel maintenances: columns Equipamento, Peça, Tipo, Valor, Data, Troca de Óleo
7. PDF combined: HTTP 200 and readable columns
8. Dates in DD/MM/AAAA format
9. Monetary values in R$ X.XXX,XX format
10. Excel produtos_admin: columns Código, Descrição, Unidade, Preço, Estoque
"""

import pytest
import requests
import os
import io
import re
import pdfplumber
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
TEST_EMAIL = "test@test.com"
TEST_PASSWORD = "password"


class TestExportFormats:
    """Test export format standardization for PDF and Excel"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup: login and get token"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": TEST_EMAIL,
            "password": TEST_PASSWORD
        })
        assert response.status_code == 200, f"Login failed: {response.text}"
        data = response.json()
        self.token = data.get("token")
        assert self.token, "Token not found in login response"
        self.headers = {"Authorization": f"Bearer {self.token}"}
    
    # ============ PDF TESTS ============
    
    def test_pdf_contas_pagar_columns_and_format(self):
        """Test 1: PDF contas_pagar has correct columns and formats"""
        response = requests.get(
            f"{BASE_URL}/api/export/pdf/contas_pagar",
            headers=self.headers
        )
        assert response.status_code == 200, f"PDF export failed: {response.status_code}"
        assert "application/pdf" in response.headers.get("Content-Type", "")
        
        # Parse PDF content
        pdf_buffer = io.BytesIO(response.content)
        with pdfplumber.open(pdf_buffer) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""
        
        # Check required column headers
        required_headers = ["Descrição", "Valor", "Vencimento", "Quitação", "Status", "Fornecedor"]
        for header in required_headers:
            assert header in text, f"Missing column header '{header}' in PDF contas_pagar"
        
        # Check that raw field names are NOT present
        raw_fields = ["fornecedor_id", "data_vencimento", "data_pagamento", "_id"]
        for field in raw_fields:
            assert field not in text, f"Raw field '{field}' should not appear in PDF"
        
        print(f"✓ PDF contas_pagar has all required columns: {required_headers}")
    
    def test_pdf_contas_receber_columns_and_format(self):
        """Test 2: PDF contas_receber has correct columns and formats"""
        response = requests.get(
            f"{BASE_URL}/api/export/pdf/contas_receber",
            headers=self.headers
        )
        assert response.status_code == 200, f"PDF export failed: {response.status_code}"
        
        pdf_buffer = io.BytesIO(response.content)
        with pdfplumber.open(pdf_buffer) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""
        
        # Check required column headers
        required_headers = ["Descrição", "Valor", "Vencimento", "Recebimento", "Status", "Cliente"]
        for header in required_headers:
            assert header in text, f"Missing column header '{header}' in PDF contas_receber"
        
        # Check that raw field names are NOT present
        raw_fields = ["cliente_id", "data_vencimento", "data_recebimento", "_id"]
        for field in raw_fields:
            assert field not in text, f"Raw field '{field}' should not appear in PDF"
        
        print(f"✓ PDF contas_receber has all required columns: {required_headers}")
    
    def test_pdf_date_format_dd_mm_yyyy(self):
        """Test 8: Dates in PDF should be in DD/MM/AAAA format, not YYYY-MM-DD"""
        response = requests.get(
            f"{BASE_URL}/api/export/pdf/contas_pagar",
            headers=self.headers
        )
        assert response.status_code == 200
        
        pdf_buffer = io.BytesIO(response.content)
        with pdfplumber.open(pdf_buffer) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""
        
        # Check for YYYY-MM-DD format (should NOT be present in data rows)
        # Exclude the filename which may have YYYYMMDD
        lines = text.split('\n')
        for line in lines:
            # Skip header lines and empty lines
            if not line.strip() or "Gerado em" in line or "Relatório" in line:
                continue
            # Check if line contains YYYY-MM-DD format (ISO date)
            iso_dates = re.findall(r'\b\d{4}-\d{2}-\d{2}\b', line)
            assert len(iso_dates) == 0, f"Found ISO date format in PDF: {iso_dates} in line: {line}"
        
        # Check for DD/MM/YYYY format (should be present if there's data)
        br_dates = re.findall(r'\b\d{2}/\d{2}/\d{4}\b', text)
        print(f"✓ PDF dates are in DD/MM/AAAA format. Found {len(br_dates)} dates.")
    
    def test_pdf_money_format_brazilian(self):
        """Test 9: Monetary values in PDF should be in R$ X.XXX,XX format"""
        response = requests.get(
            f"{BASE_URL}/api/export/pdf/contas_pagar",
            headers=self.headers
        )
        assert response.status_code == 200
        
        pdf_buffer = io.BytesIO(response.content)
        with pdfplumber.open(pdf_buffer) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""
        
        # Check for R$ format (Brazilian currency)
        money_pattern = re.findall(r'R\$\s*[\d.,]+', text)
        if money_pattern:
            print(f"✓ Found Brazilian currency format: {money_pattern[:5]}...")
            # Verify format uses comma as decimal separator
            for money in money_pattern:
                # Should have comma before last 2 digits (decimal separator)
                assert ',' in money, f"Money format should use comma as decimal: {money}"
        else:
            print("⚠ No monetary values found in PDF (may be empty data)")
    
    # ============ EXCEL TESTS ============
    
    def test_excel_contas_pagar_columns(self):
        """Test 3: Excel contas_pagar has correct columns without raw field names"""
        response = requests.get(
            f"{BASE_URL}/api/export/excel/contas_pagar",
            headers=self.headers
        )
        assert response.status_code == 200, f"Excel export failed: {response.status_code}"
        
        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1] if cell.value]
        
        # Required columns
        required = ["Descrição", "Valor", "Vencimento", "Quitação", "Status", "Fornecedor", "Centro de Custo", "Plano de Contas"]
        for col in required:
            assert col in headers, f"Missing column '{col}' in Excel contas_pagar. Found: {headers}"
        
        # Should NOT have raw field names
        forbidden = ["fornecedor_id", "data_vencimento", "data_pagamento", "_id", "id"]
        for col in forbidden:
            assert col not in headers, f"Raw field '{col}' should not be in Excel headers"
        
        print(f"✓ Excel contas_pagar has correct columns: {headers}")
    
    def test_excel_contas_receber_columns(self):
        """Test 4: Excel contas_receber has correct columns"""
        response = requests.get(
            f"{BASE_URL}/api/export/excel/contas_receber",
            headers=self.headers
        )
        assert response.status_code == 200, f"Excel export failed: {response.status_code}"
        
        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1] if cell.value]
        
        required = ["Descrição", "Valor", "Vencimento", "Recebimento", "Status", "Cliente", "Centro de Custo", "Plano de Contas"]
        for col in required:
            assert col in headers, f"Missing column '{col}' in Excel contas_receber. Found: {headers}"
        
        forbidden = ["cliente_id", "data_vencimento", "data_recebimento", "_id", "id"]
        for col in forbidden:
            assert col not in headers, f"Raw field '{col}' should not be in Excel headers"
        
        print(f"✓ Excel contas_receber has correct columns: {headers}")
    
    def test_excel_obras_columns(self):
        """Test 5: Excel obras has correct columns without raw field names"""
        response = requests.get(
            f"{BASE_URL}/api/export/excel/obras",
            headers=self.headers
        )
        assert response.status_code == 200, f"Excel export failed: {response.status_code}"
        
        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1] if cell.value]
        
        required = ["Nome", "Local", "Cliente", "Status", "Data Início", "Data Fim", "Contrato"]
        for col in required:
            assert col in headers, f"Missing column '{col}' in Excel obras. Found: {headers}"
        
        forbidden = ["cliente_id", "start_date", "end_date", "_id", "id"]
        for col in forbidden:
            assert col not in headers, f"Raw field '{col}' should not be in Excel headers"
        
        print(f"✓ Excel obras has correct columns: {headers}")
    
    def test_excel_maintenances_columns(self):
        """Test 6: Excel maintenances has correct columns"""
        response = requests.get(
            f"{BASE_URL}/api/export/excel/maintenances",
            headers=self.headers
        )
        assert response.status_code == 200, f"Excel export failed: {response.status_code}"
        
        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1] if cell.value]
        
        required = ["Equipamento", "Peça", "Tipo", "Valor", "Data", "Troca de Óleo"]
        for col in required:
            assert col in headers, f"Missing column '{col}' in Excel maintenances. Found: {headers}"
        
        forbidden = ["machine_id", "replacement_date", "_id", "id"]
        for col in forbidden:
            assert col not in headers, f"Raw field '{col}' should not be in Excel headers"
        
        print(f"✓ Excel maintenances has correct columns: {headers}")
    
    def test_excel_produtos_admin_columns(self):
        """Test 10: Excel produtos_admin has correct columns"""
        response = requests.get(
            f"{BASE_URL}/api/export/excel/produtos_admin",
            headers=self.headers
        )
        assert response.status_code == 200, f"Excel export failed: {response.status_code}"
        
        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1] if cell.value]
        
        required = ["Código", "Descrição", "Unidade", "Preço", "Estoque"]
        for col in required:
            assert col in headers, f"Missing column '{col}' in Excel produtos_admin. Found: {headers}"
        
        print(f"✓ Excel produtos_admin has correct columns: {headers}")
    
    def test_excel_date_format_dd_mm_yyyy(self):
        """Test 8 (Excel): Dates should be in DD/MM/AAAA format"""
        response = requests.get(
            f"{BASE_URL}/api/export/excel/contas_pagar",
            headers=self.headers
        )
        assert response.status_code == 200
        
        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Check date columns (Vencimento, Quitação - columns C and D)
        date_values = []
        for row in ws.iter_rows(min_row=2, max_row=min(10, ws.max_row)):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Check if it's a date string
                    if re.match(r'\d{2}/\d{2}/\d{4}', str(cell.value)):
                        date_values.append(cell.value)
                    # Should NOT have YYYY-MM-DD format
                    if re.match(r'\d{4}-\d{2}-\d{2}', str(cell.value)):
                        pytest.fail(f"Found ISO date format in Excel: {cell.value}")
        
        if date_values:
            print(f"✓ Excel dates are in DD/MM/AAAA format: {date_values[:3]}...")
        else:
            print("⚠ No date values found in Excel (may be empty data or using date format)")
    
    # ============ COMBINED PDF TEST ============
    
    def test_pdf_combined_export(self):
        """Test 7: PDF combined export returns HTTP 200 and has readable columns"""
        response = requests.post(
            f"{BASE_URL}/api/export/combined",
            headers=self.headers,
            json={
                "categories": ["contas_pagar", "contas_receber"],
                "format": "pdf"
            }
        )
        assert response.status_code == 200, f"Combined PDF export failed: {response.status_code} - {response.text}"
        assert "application/pdf" in response.headers.get("Content-Type", "")
        
        pdf_buffer = io.BytesIO(response.content)
        with pdfplumber.open(pdf_buffer) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""
        
        # Should have readable column headers, not raw field names
        readable_headers = ["Descrição", "Valor", "Vencimento", "Status"]
        for header in readable_headers:
            assert header in text, f"Missing readable header '{header}' in combined PDF"
        
        # Should NOT have raw field names
        raw_fields = ["fornecedor_id", "cliente_id", "data_vencimento", "_id"]
        for field in raw_fields:
            assert field not in text, f"Raw field '{field}' should not appear in combined PDF"
        
        print("✓ Combined PDF export works with readable columns")
    
    def test_pdf_combined_with_centro_custo(self):
        """Test combined PDF with centro_custo filter"""
        response = requests.post(
            f"{BASE_URL}/api/export/combined",
            headers=self.headers,
            json={
                "categories": ["contas_pagar", "contas_receber"],
                "format": "pdf",
                "centro_custo": "Administrativo"
            }
        )
        assert response.status_code == 200, f"Combined PDF with CC failed: {response.status_code}"
        print("✓ Combined PDF with centro_custo filter works")


class TestExportEndpointsStatus:
    """Test that all export endpoints return HTTP 200"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup: login and get token"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": TEST_EMAIL,
            "password": TEST_PASSWORD
        })
        assert response.status_code == 200
        self.token = response.json().get("token")
        self.headers = {"Authorization": f"Bearer {self.token}"}
    
    @pytest.mark.parametrize("category", [
        "contas_pagar",
        "contas_receber",
        "obras",
        "maintenances",
        "machines",
        "stock_items",
        "cadastros"
    ])
    def test_pdf_export_status(self, category):
        """Test PDF export returns 200 for various categories"""
        response = requests.get(
            f"{BASE_URL}/api/export/pdf/{category}",
            headers=self.headers
        )
        assert response.status_code == 200, f"PDF {category} failed: {response.status_code}"
        print(f"✓ PDF {category}: HTTP 200")
    
    @pytest.mark.parametrize("category", [
        "contas_pagar",
        "contas_receber",
        "obras",
        "maintenances",
        "machines",
        "stock_items",
        "cadastros",
        "produtos_admin"
    ])
    def test_excel_export_status(self, category):
        """Test Excel export returns 200 for various categories"""
        response = requests.get(
            f"{BASE_URL}/api/export/excel/{category}",
            headers=self.headers
        )
        assert response.status_code == 200, f"Excel {category} failed: {response.status_code}"
        print(f"✓ Excel {category}: HTTP 200")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
